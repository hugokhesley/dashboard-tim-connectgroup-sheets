import streamlit as st
import pandas as pd
from datetime import datetime
import unicodedata
import re
from data_loader import get_gspread_client, _s, _to_num, _normalize, _norm_pedido, _dedup_columns, get_meta_mes, registrar_acesso, load_bko, load_colaboradores, load_data, apply_filters, get_parceiros
from auth import require_login
from ui import aplicar_estilo_base

st.set_page_config(
    page_title="Connect Group | Resultados",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

username = require_login("resultados")
aplicar_estilo_base()
registrar_acesso("resultados", username=username)

st.markdown("""
<style>
  .header-res {
    background: linear-gradient(135deg, #1e1b4b 0%, #3730a3 50%, #6366f1 100%);
    border-radius: 16px; padding: 28px 36px; margin-bottom: 28px;
    box-shadow: 0 8px 32px rgba(99,102,241,0.3);
    border: 1px solid rgba(255,255,255,0.08);
    display: flex; align-items: center; justify-content: space-between;
  }
  .header-title { font-size: 1.9rem; font-weight: 800; color: #fff; margin: 0; }
  .header-sub   { font-size: 0.85rem; color: rgba(255,255,255,0.65); margin: 4px 0 0 0; }
  .header-badge {
    background: rgba(255,255,255,0.15); border: 1px solid rgba(255,255,255,0.25);
    border-radius: 20px; padding: 6px 16px; font-size: 0.8rem; color: #fff; font-weight: 600;
  }
  .kpi-card {
    background: #1a1f2e; border-radius: 14px; padding: 22px 24px;
    border: 1px solid #2d3748; position: relative; overflow: hidden; margin-bottom: 4px;
  }
  .kpi-card::before { content:''; position:absolute; top:0; left:0; right:0; height:3px; }
  .kpi-card.indigo::before { background: linear-gradient(90deg, #6366f1, #4f46e5); }
  .kpi-card.green::before  { background: linear-gradient(90deg, #10b981, #059669); }
  .kpi-card.amber::before  { background: linear-gradient(90deg, #f59e0b, #d97706); }
  .kpi-card.purple::before { background: linear-gradient(90deg, #8b5cf6, #7c3aed); }
  .kpi-label { font-size: 0.72rem; text-transform: uppercase; letter-spacing: 1px; color: #94a3b8; font-weight: 600; margin-bottom: 8px; }
  .kpi-value { font-size: 2.1rem; font-weight: 800; color: #f1f5f9; line-height: 1; }
  .kpi-sub   { font-size: 0.78rem; color: #64748b; margin-top: 6px; }
  section[data-testid="stSidebar"] { background: #111827 !important; }
  .header-logo { height:44px;width:auto;object-fit:contain;mix-blend-mode:multiply;border-radius:6px; }
  .header-right { display:flex;align-items:center;gap:14px; }
  .section-title { font-size:0.75rem; text-transform:uppercase; letter-spacing:1.5px; color:#64748b; font-weight:600; margin:24px 0 12px 0; }
  .rank-bar-wrap { background:#1a1f2e; border-radius:10px; padding:12px 16px; margin-bottom:8px; border:1px solid #2d3748; }
  .rank-name     { display:flex; justify-content:space-between; font-size:0.82rem; font-weight:600; color:#e2e8f0; margin-bottom:6px; }
  .rank-bar-bg   { background:#2d3748; border-radius:99px; height:8px; }
  .rank-bar-fill { height:8px; border-radius:99px; }
  .rank-val      { font-size:0.75rem; color:#64748b; margin-top:4px; }
  .equipe-header { background:linear-gradient(90deg,#1a2e1a,#1e3a2e); border-left:4px solid #22c55e; border-radius:10px; padding:14px 20px; margin:20px 0 12px 0; display:flex; align-items:center; justify-content:space-between; }
  .equipe-nome   { font-size:1rem; font-weight:800; color:#86efac; }
  .equipe-total  { font-size:0.82rem; color:#64748b; }
</style>
""", unsafe_allow_html=True)


@st.cache_data(ttl=180)
def load_resultados():
    try:
        client = get_gspread_client()
        sheet_url = st.secrets["sheets"]["url"]
        spreadsheet = client.open_by_url(sheet_url)
        ws = spreadsheet.worksheet("resultados")
        all_values = ws.get_all_values()
        if not all_values or len(all_values) < 2:
            return pd.DataFrame()
        headers = all_values[0]
        rows = all_values[1:]
        df = pd.DataFrame(rows, columns=headers)
        df = _dedup_columns(df)
        df.columns = [_s(c).lower() for c in df.columns]
        df = _dedup_columns(df)
        return df
    except Exception as e:
        st.error(f"Erro ao carregar aba resultados: {e}")
        return pd.DataFrame()


def normalize_resultados(df):
    df = _dedup_columns(df)
    rename = {}
    for col in df.columns:
        n = _normalize(col)
        if n == "data de ativacao":       rename[col] = "data_ativacao"
        elif n == "razao social":         rename[col] = "razao_social"
        elif n == "tipo de contratacao":  rename[col] = "tipo_contratacao"
        elif n == "acessos":              rename[col] = "acessos"
        elif n == "preco oferta":         rename[col] = "preco_oferta"
        elif n == "parceiro":             rename[col] = "parceiro"
        elif n == "fila atual":           rename[col] = "fila_atual"
        elif n == "pedido":               rename[col] = "pedido"
        elif n == "cnpj":                 rename[col] = "cnpj"
    df = df.rename(columns=rename)
    df = _dedup_columns(df)
    for col in ["razao_social", "tipo_contratacao", "parceiro"]:
        if col in df.columns:
            df[col] = df[col].apply(_s)
    if "acessos" in df.columns:
        df["acessos"] = df["acessos"].apply(_to_num)
    if "preco_oferta" in df.columns:
        df["preco_oferta"] = df["preco_oferta"].apply(_to_num)
    # FIX 4 — normaliza o pedido com a MESMA regra do BKO (_norm_pedido),
    # senão o merge on="pedido" não casa quando vem "6342770.0" ou com espaço.
    if "pedido" in df.columns:
        df["pedido"] = df["pedido"].apply(_norm_pedido)
    if "data_ativacao" in df.columns:
        df["mes_ativacao"] = pd.to_datetime(
            df["data_ativacao"].apply(_s), dayfirst=True, errors="coerce"
        ).dt.strftime("%m/%Y")
    return df


def _prog(v, t, color):
    pct = min(int(v / t * 100), 100) if t > 0 else 0
    return (
        '<div style="margin-top:10px">'
        '<div style="display:flex;justify-content:space-between;font-size:0.71rem;color:#64748b;margin-bottom:4px">'
        f'<span>Atingimento</span><span>{pct}%</span></div>'
        '<div style="background:#2d3748;border-radius:99px;height:7px">'
        f'<div style="width:{pct}%;background:{color};height:7px;border-radius:99px"></div>'
        '</div></div>'
    )




META_VENDEDOR_PAD = 850

def _meta_vend(nome, meta_dict):
    return meta_dict.get(nome, META_VENDEDOR_PAD)

def _cor(pct):
    if pct >= 100: return "#22c55e"
    if pct >= 70:  return "#f59e0b"
    return "#ef4444"

def _bar(valor, maximo, cor="#22c55e", h=8):
    pct = min(int(valor / maximo * 100), 100) if maximo > 0 else 0
    return f'''<div class="rank-bar-bg"><div class="rank-bar-fill" style="width:{pct}%;background:{cor};height:{h}px"></div></div>'''

def render_ranking_resultados(df_atv, mes_alvo, meta_dict, lideres):
    """Ranking de vendedores e equipes com base nos ativados do mês."""
    st.markdown('<p class="section-title">🏆 Ranking por Vendedor Real</p>', unsafe_allow_html=True)

    if "vendedor_real" not in df_atv.columns or df_atv.empty:
        st.info("BKO não carregado ou sem dados de vendedor para este mês.")
        return

    # ── Ranking Geral de Vendedores ───────────────────────────────────────
    rank_v = (df_atv.groupby("vendedor_real", as_index=False)
              .agg(Acessos=("acessos","sum"), Receita=("preco_oferta","sum"))
              .sort_values("Receita", ascending=False))

    # Renderiza ranking em grid 2 colunas com expander nativo para clientes
    cols_rank = st.columns(2)
    for i, (_, r) in enumerate(rank_v.iterrows()):
        meta  = _meta_vend(r["vendedor_real"], meta_dict)
        pct   = min(int(r["Receita"] / meta * 100), 100) if meta > 0 else 0
        cor   = _cor(pct)
        medal = ["🥇","🥈","🥉"][i] if i < 3 else f"{i+1}º"

        with cols_rank[i % 2]:
            # Card do vendedor
            st.markdown(f'''<div class="rank-bar-wrap">
              <div class="rank-name">
                <span>{medal} {r["vendedor_real"]}</span>
                <span style="color:{cor};font-weight:700">{pct}%</span>
              </div>
              {_bar(r["Receita"], meta, cor)}
              <div class="rank-val">R$ {r["Receita"]:,.2f} · {int(r["Acessos"])} acessos · meta R$ {meta:,.2f}</div>
            </div>''', unsafe_allow_html=True)

            # Expander com clientes ativados
            clientes_v = df_atv[df_atv["vendedor_real"] == r["vendedor_real"]]
            if "razao_social" in clientes_v.columns and not clientes_v.empty:
                with st.expander("👁 Ver clientes ativados"):
                    top_cli = (clientes_v.groupby("razao_social")["acessos"]
                               .sum().sort_values(ascending=False)
                               .reset_index())
                    top_cli.columns = ["Cliente", "Linhas"]
                    top_cli["Linhas"] = top_cli["Linhas"].astype(int)
                    st.dataframe(top_cli, use_container_width=True, hide_index=True,
                        column_config={
                            "Cliente": st.column_config.TextColumn("Cliente"),
                            "Linhas":  st.column_config.NumberColumn("Linhas", format="%d"),
                        })

    # ── Ranking por Equipe ────────────────────────────────────────────────
    if "lider" in df_atv.columns and df_atv["lider"].nunique() > 1:
        st.markdown('<p class="section-title">👥 Ranking por Equipe</p>', unsafe_allow_html=True)

        rows = []
        for lider in lideres:
            dl = df_atv[df_atv["lider"] == lider]
            if dl.empty: continue
            rec    = dl["preco_oferta"].sum()
            ac     = int(dl["acessos"].sum())
            meta_e = sum(_meta_vend(v, meta_dict) for v in dl["vendedor_real"].unique())
            pct_e  = min(int(rec / meta_e * 100), 100) if meta_e > 0 else 0
            rows.append({"lider": lider, "rec": rec, "ac": ac, "meta": meta_e, "pct": pct_e})

        rows = sorted(rows, key=lambda x: x["rec"], reverse=True)
        html_e = ""
        for r in rows:
            cor_e = _cor(r["pct"])
            icon  = "✅" if r["pct"] >= 100 else "⚠️" if r["pct"] >= 70 else "🔴"
            html_e += f'''<div class="equipe-header">
              <span class="equipe-nome">👤 {r["lider"]}</span>
              <span class="equipe-total">{icon} {r["pct"]}% da meta · R$ {r["rec"]:,.2f} · {r["ac"]} acessos</span>
            </div>'''

            dl_eq = df_atv[df_atv["lider"] == r["lider"]]
            rank_eq = (dl_eq.groupby("vendedor_real", as_index=False)
                       .agg(Acessos=("acessos","sum"), Receita=("preco_oferta","sum"))
                       .sort_values("Receita", ascending=False))
            html_eq = ''
            for j, rv in rank_eq.iterrows():
                meta_v = _meta_vend(rv["vendedor_real"], meta_dict)
                pct_v  = min(int(rv["Receita"] / meta_v * 100), 100) if meta_v > 0 else 0
                cor_v  = _cor(pct_v)
                html_eq += f'''<div class="rank-bar-wrap" style="margin-left:16px">
                  <div class="rank-name">
                    <span>{rv["vendedor_real"]}</span>
                    <span style="color:{cor_v}">{pct_v}%</span>
                  </div>
                  {_bar(rv["Receita"], meta_v, cor_v)}
                  <div class="rank-val">R$ {rv["Receita"]:,.2f} · {int(rv["Acessos"])} acessos</div>
                </div>'''
            html_e += html_eq

        st.markdown(html_e, unsafe_allow_html=True)

def gerar_pdf_resultados(df_atv, meta_dict, mes):
    """Gera PDF formatado Equipe → Vendedor → Clientes."""
    import io
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_RIGHT, TA_CENTER

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
        leftMargin=14*mm, rightMargin=14*mm,
        topMargin=14*mm, bottomMargin=14*mm)

    VERDE     = colors.HexColor("#22c55e")
    VERDE_ESC = colors.HexColor("#15803d")
    AZUL      = colors.HexColor("#3b82f6")
    AMBER     = colors.HexColor("#f59e0b")
    CINZA     = colors.HexColor("#64748b")
    CINZA_ESC = colors.HexColor("#1e293b")
    BRANCO    = colors.white
    FL        = colors.HexColor("#131f2e")
    FV        = colors.HexColor("#0f1820")
    FC        = colors.HexColor("#080e15")
    TEXTO     = colors.HexColor("#e2e8f0")
    TDIM      = colors.HexColor("#94a3b8")

    def s(name, **kw):
        base = {"fontName": "Helvetica", "fontSize": 9, "textColor": TEXTO, "leading": 13}
        base.update(kw)
        return ParagraphStyle(name, **base)

    sTitle  = s("T", fontSize=15, fontName="Helvetica-Bold", textColor=BRANCO)
    sRight  = s("R", fontSize=7,  textColor=BRANCO, alignment=TA_RIGHT)
    sLider  = s("L", fontSize=9,  fontName="Helvetica-Bold", textColor=BRANCO)
    sVend   = s("V", fontSize=8,  fontName="Helvetica-Bold", textColor=TEXTO)
    sCli    = s("C", fontSize=7,  textColor=TDIM)

    story = []
    agora = datetime.now().strftime("%d/%m/%Y %H:%M")

    # Header
    h = Table([[Paragraph(f"<b>CONNECT GROUP — Vendas {mes}</b>", sTitle),
                Paragraph(f"Gerado em {agora}", sRight)]], colWidths=["70%","30%"])
    h.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),VERDE_ESC),("ROWPADDING",(0,0),(-1,-1),10),("VALIGN",(0,0),(-1,-1),"MIDDLE")]))
    story.append(h)
    story.append(Spacer(1, 6*mm))

    lideres = sorted([l for l in df_atv["lider"].dropna().unique() if l and l != "Sem Equipe"])

    for lider in lideres:
        df_l  = df_atv[df_atv["lider"] == lider]
        ac_l  = int(df_l["acessos"].sum())
        rec_l = df_l["preco_oferta"].sum()
        vends = sorted([v for v in df_l["vendedor_real"].dropna().unique() if v and v != "Sem Vendedor"])

        lr = Table([[Paragraph(f"  {lider}", sLider),
                     Paragraph(f"{len(vends)} vend.   {ac_l} ac / R$ {rec_l:,.2f}", sRight)]],
                   colWidths=["55%","45%"])
        lr.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),AZUL),("ROWPADDING",(0,0),(-1,-1),8),
                                 ("VALIGN",(0,0),(-1,-1),"MIDDLE"),("LINEBELOW",(0,0),(-1,0),2,colors.HexColor("#1d4ed8"))]))
        story.append(lr)

        for vend in vends:
            df_v   = df_l[df_l["vendedor_real"] == vend]
            ac_v   = int(df_v["acessos"].sum())
            rec_v  = df_v["preco_oferta"].sum()
            meta_v = _meta_vend(vend, meta_dict)
            pct_v  = min(int(rec_v / meta_v * 100), 100) if meta_v > 0 else 0
            cor_v  = VERDE if pct_v >= 100 else AMBER if pct_v >= 70 else colors.HexColor("#ef4444")

            vr = Table([[Paragraph(f"    {vend}", sVend),
                         Paragraph(f"{ac_v} ac / R$ {rec_v:,.2f} ({pct_v}%)", sRight)]],
                       colWidths=["55%","45%"])
            vr.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),FL),("ROWPADDING",(0,0),(-1,-1),6),
                                     ("VALIGN",(0,0),(-1,-1),"MIDDLE"),("LINEBELOW",(0,0),(-1,0),0.5,CINZA_ESC),
                                     ("TEXTCOLOR",(1,0),(1,0),cor_v)]))
            story.append(vr)

            # Clientes
            cols_grp = [c for c in ["razao_social","fila_atual"] if c in df_v.columns]
            if cols_grp:
                df_g = df_v.copy()
                for col in cols_grp: df_g[col] = df_g[col].fillna("—")
                cdf = (df_g.groupby(cols_grp, as_index=False)
                       .agg(ac=("acessos","sum"), rec=("preco_oferta","sum"))
                       .sort_values("ac", ascending=False))
                for _, crow in cdf.iterrows():
                    razao = str(crow.get("razao_social","—"))[:55]
                    fila  = str(crow.get("fila_atual","—")).upper()
                    ac_c  = int(crow.get("ac",0))
                    rec_c = float(crow.get("rec",0))
                    cr = Table([[Paragraph(f"        {razao}", sCli),
                                 Paragraph(f"{ac_c} ac   R$ {rec_c:,.2f}   {fila}", sRight)]],
                               colWidths=["60%","40%"])
                    cr.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),FC),("ROWPADDING",(0,0),(-1,-1),3),
                                             ("VALIGN",(0,0),(-1,-1),"MIDDLE"),("LINEBELOW",(0,0),(-1,0),0.3,CINZA_ESC)]))
                    story.append(cr)

        story.append(Spacer(1, 4*mm))

    story.append(HRFlowable(width="100%", thickness=0.5, color=CINZA_ESC))
    story.append(Spacer(1, 1*mm))
    story.append(Paragraph(f"Connect Group · {mes} · {agora}",
                            s("F", fontSize=6, textColor=CINZA, alignment=TA_CENTER)))
    doc.build(story)
    buf.seek(0)
    return buf.read()


def gerar_excel_resultados(df_atv, meta_dict, mes):
    """Gera Excel formatado Equipe → Vendedor → Clientes."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = f"Vendas {mes.replace('/', '-')}"

    # Estilos
    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color.replace("#",""))

    def border():
        thin = Side(style="thin", color="1E293B")
        return Border(bottom=thin)

    ft_title  = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ft_lider  = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ft_vend   = Font(name="Calibri", bold=True, size=10, color="E2E8F0")
    ft_cli    = Font(name="Calibri", size=9,    color="94A3B8")
    ft_header = Font(name="Calibri", bold=True, size=9,  color="FFFFFF")

    al_left  = Alignment(horizontal="left",  vertical="center")
    al_right = Alignment(horizontal="right", vertical="center")
    al_center= Alignment(horizontal="center",vertical="center")

    agora = datetime.now().strftime("%d/%m/%Y %H:%M")

    # Larguras
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 20

    row = 1
    # Título
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = f"CONNECT GROUP — Vendas {mes}"
    ws[f"A{row}"].font = ft_title
    ws[f"A{row}"].fill = fill("#15803d")
    ws[f"A{row}"].alignment = al_left
    ws.row_dimensions[row].height = 28
    row += 1

    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = f"Gerado em {agora}"
    ws[f"A{row}"].font = Font(name="Calibri", size=9, color="94A3B8")
    ws[f"A{row}"].fill = fill("#0f1117")
    ws[f"A{row}"].alignment = al_left
    row += 2

    lideres = sorted([l for l in df_atv["lider"].dropna().unique() if l and l != "Sem Equipe"])

    for lider in lideres:
        df_l  = df_atv[df_atv["lider"] == lider]
        ac_l  = int(df_l["acessos"].sum())
        rec_l = df_l["preco_oferta"].sum()
        vends = sorted([v for v in df_l["vendedor_real"].dropna().unique() if v and v != "Sem Vendedor"])

        # Linha líder
        ws.merge_cells(f"A{row}:C{row}")
        ws[f"A{row}"] = f"  {lider}"
        ws[f"A{row}"].font = ft_lider
        ws[f"A{row}"].fill = fill("#1d4ed8")
        ws[f"A{row}"].alignment = al_left
        ws[f"D{row}"] = ac_l
        ws[f"D{row}"].font = ft_lider
        ws[f"D{row}"].fill = fill("#1d4ed8")
        ws[f"D{row}"].alignment = al_right
        ws[f"E{row}"] = rec_l
        ws[f"E{row}"].font = ft_lider
        ws[f"E{row}"].fill = fill("#1d4ed8")
        ws[f"E{row}"].alignment = al_right
        ws[f"E{row}"].number_format = "R$ #,##0.00"
        ws.row_dimensions[row].height = 20
        row += 1

        for vend in vends:
            df_v   = df_l[df_l["vendedor_real"] == vend]
            ac_v   = int(df_v["acessos"].sum())
            rec_v  = df_v["preco_oferta"].sum()
            meta_v = _meta_vend(vend, meta_dict)
            pct_v  = min(int(rec_v / meta_v * 100), 100) if meta_v > 0 else 0
            cor_v  = "22C55E" if pct_v >= 100 else "F59E0B" if pct_v >= 70 else "EF4444"

            ws.merge_cells(f"A{row}:C{row}")
            ws[f"A{row}"] = f"    {vend}"
            ws[f"A{row}"].font = ft_vend
            ws[f"A{row}"].fill = fill("#131f2e")
            ws[f"A{row}"].alignment = al_left
            ws[f"D{row}"] = ac_v
            ws[f"D{row}"].font = Font(name="Calibri", bold=True, size=10, color=cor_v)
            ws[f"D{row}"].fill = fill("#131f2e")
            ws[f"D{row}"].alignment = al_right
            ws[f"E{row}"] = rec_v
            ws[f"E{row}"].font = Font(name="Calibri", bold=True, size=10, color=cor_v)
            ws[f"E{row}"].fill = fill("#131f2e")
            ws[f"E{row}"].alignment = al_right
            ws[f"E{row}"].number_format = "R$ #,##0.00"
            ws.row_dimensions[row].height = 18
            row += 1

            # Clientes
            cols_grp = [c for c in ["razao_social","fila_atual"] if c in df_v.columns]
            if cols_grp:
                df_g = df_v.copy()
                for col in cols_grp: df_g[col] = df_g[col].fillna("—")
                cdf = (df_g.groupby(cols_grp, as_index=False)
                       .agg(ac=("acessos","sum"), rec=("preco_oferta","sum"))
                       .sort_values("ac", ascending=False))
                for _, crow in cdf.iterrows():
                    razao = str(crow.get("razao_social","—"))[:60]
                    fila  = str(crow.get("fila_atual","—")).upper()
                    ac_c  = int(crow.get("ac",0))
                    rec_c = float(crow.get("rec",0))
                    ws.merge_cells(f"A{row}:B{row}")
                    ws[f"A{row}"] = f"        {razao}"
                    ws[f"A{row}"].font = ft_cli
                    ws[f"A{row}"].fill = fill("#080e15")
                    ws[f"A{row}"].alignment = al_left
                    ws[f"C{row}"] = fila
                    ws[f"C{row}"].font = ft_cli
                    ws[f"C{row}"].fill = fill("#080e15")
                    ws[f"C{row}"].alignment = al_center
                    ws[f"D{row}"] = ac_c
                    ws[f"D{row}"].font = ft_cli
                    ws[f"D{row}"].fill = fill("#080e15")
                    ws[f"D{row}"].alignment = al_right
                    ws[f"E{row}"] = rec_c
                    ws[f"E{row}"].font = ft_cli
                    ws[f"E{row}"].fill = fill("#080e15")
                    ws[f"E{row}"].alignment = al_right
                    ws[f"E{row}"].number_format = "R$ #,##0.00"
                    ws.row_dimensions[row].height = 15
                    row += 1

        row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def main():
    st.markdown("""
    <div class="header-res">
      <div>
        <p class="header-title">📊 RESULTADOS — CONNECT GROUP</p>
        <p class="header-sub">TIM Corporate · Análise Mensal de Ativações</p>
      </div>
      <div class="header-right">
        <img src="https://raw.githubusercontent.com/hugokhesley/dashboard-tim-connectgroup-sheets/main/logo.png" class="header-logo" onerror="this.style.display='none'">
        <span class="header-badge">📈 RESULTADOS</span>
      </div>
    </div>""", unsafe_allow_html=True)

    with st.spinner("Carregando aba resultados..."):
        raw = load_resultados()

    if raw.empty:
        st.warning("Nenhum dado encontrado na aba 'resultados'. Verifique se a aba existe na planilha.")
        st.stop()

    df = normalize_resultados(raw.copy())

    # ── Sidebar filtros ──────────────────────────────────────────────────
    with st.sidebar:
        st.markdown("### 🔧 Filtros")

        if "tipo_contratacao" in df.columns:
            tipos_raw = sorted([t for t in df["tipo_contratacao"].apply(
                lambda x: _s(x).upper()).unique().tolist() if t])
        else:
            tipos_raw = ["NOVO", "ADITIVO", "RENEGOCIACAO"]

        tipos_sel = st.multiselect("Tipo de Contratação", options=tipos_raw, default=tipos_raw)

        parceiros = ["Todos"]
        if "parceiro" in df.columns:
            parceiros += sorted([_s(v) for v in df["parceiro"].unique() if _s(v)])
        parceiro_sel = st.selectbox("Parceiro", parceiros)

        meses = ["Todos"]
        if "mes_ativacao" in df.columns:
            meses_validos = sorted([m for m in df["mes_ativacao"].dropna().unique() if m and m != "NaT"])
            meses += meses_validos
        mes_sel = st.selectbox("Mês de Ativação", meses, index=len(meses)-1 if len(meses) > 1 else 0)

        # Multiselect de líderes — carregado do colab
        from data_loader import load_colaboradores as _lc
        _colab_side = _lc()
        lideres_disponiveis = sorted(_colab_side["lider"].dropna().unique().tolist()) if not _colab_side.empty else []
        lideres_sel = st.multiselect(
            "👥 Equipes / Líderes",
            options=lideres_disponiveis,
            default=lideres_disponiveis,
            placeholder="Selecione as equipes..."
        )

        st.markdown("---")
        if st.button("🔄 Atualizar dados"):
            st.cache_data.clear()
            st.rerun()
        st.caption("Dados via Google Sheets · cache 3 min")

    # ── Aplicar filtros ──────────────────────────────────────────────────
    dff = df.copy()
    if tipos_sel and "tipo_contratacao" in dff.columns:
        dff = dff[dff["tipo_contratacao"].apply(lambda x: _s(x).upper() in tipos_sel)]
    if parceiro_sel != "Todos" and "parceiro" in dff.columns:
        dff = dff[dff["parceiro"].apply(lambda x: _s(x) == parceiro_sel)]
    if mes_sel != "Todos" and "mes_ativacao" in dff.columns:
        dff = dff[dff["mes_ativacao"] == mes_sel]
    dff = _dedup_columns(dff.reset_index(drop=True))

    # ── KPIs ─────────────────────────────────────────────────────────────
    dff_vendas = dff[dff["tipo_contratacao"].apply(
        lambda x: _s(x).upper() in ["NOVO", "ADITIVO"]
    )] if "tipo_contratacao" in dff.columns else dff

    total_acessos  = int(dff["acessos"].sum())        if "acessos"      in dff.columns else 0
    acessos_vendas = int(dff_vendas["acessos"].sum()) if "acessos"      in dff_vendas.columns else 0
    total_receita  = dff["preco_oferta"].sum()         if "preco_oferta" in dff.columns else 0
    ticket_medio   = (total_receita / total_acessos)   if total_acessos > 0 else 0
    total_clientes = dff["razao_social"].nunique()     if "razao_social" in dff.columns else 0

    meta = get_meta_mes(mes_sel) if mes_sel != "Todos" else {"vendas_acessos": 0, "vendas_receita": 0}
    meta_acessos = int(meta["vendas_acessos"])
    meta_receita = meta["vendas_receita"]

    st.markdown('<p class="section-title">📈 KPIs do Período</p>', unsafe_allow_html=True)
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        sub  = f'de <b style="color:#e2e8f0">{meta_acessos:,}</b> meta' if meta_acessos > 0 else "sem meta cadastrada"
        prog = _prog(acessos_vendas, meta_acessos, "#6366f1") if meta_acessos > 0 else ""
        st.markdown(f"""<div class="kpi-card indigo">
          <div class="kpi-label">🎯 Acessos (Novo + Aditivo)</div>
          <div style="display:flex;align-items:baseline;gap:8px;margin:6px 0">
            <span class="kpi-value">{acessos_vendas:,}</span>
            <span style="font-size:0.85rem;color:#94a3b8">{sub}</span>
          </div>{prog}
        </div>""", unsafe_allow_html=True)
    with c2:
        sub2  = f'de <b style="color:#e2e8f0">R$ {meta_receita:,.2f}</b> meta' if meta_receita > 0 else "sem meta cadastrada"
        prog2 = _prog(total_receita, meta_receita, "#10b981") if meta_receita > 0 else ""
        st.markdown(f"""<div class="kpi-card green">
          <div class="kpi-label">💰 Receita Total</div>
          <div style="display:flex;align-items:baseline;gap:8px;margin:6px 0">
            <span class="kpi-value">R$ {total_receita:,.2f}</span>
            <span style="font-size:0.85rem;color:#94a3b8">{sub2}</span>
          </div>{prog2}
        </div>""", unsafe_allow_html=True)
    with c3:
        st.markdown(f"""<div class="kpi-card amber">
          <div class="kpi-label">🎫 Ticket Médio</div>
          <div class="kpi-value">R$ {ticket_medio:,.2f}</div>
          <div class="kpi-sub">por acesso</div>
        </div>""", unsafe_allow_html=True)
    with c4:
        st.markdown(f"""<div class="kpi-card purple">
          <div class="kpi-label">🏢 Clientes</div>
          <div class="kpi-value">{total_clientes:,}</div>
          <div class="kpi-sub">razões sociais únicas</div>
        </div>""", unsafe_allow_html=True)

    # ── Busca de Cliente ──────────────────────────────────────────────────
    st.markdown('<p class="section-title">🔎 Buscar Cliente</p>', unsafe_allow_html=True)
    termo = st.text_input(
        "Buscar por razão social, CNPJ ou nº do pedido (radar)",
        placeholder="Ex.: TELEFIBRA · 12.345.678/0001-90 · 6342770",
        label_visibility="collapsed",
        key="busca_cliente",
    )

    if termo and termo.strip():
        termo_txt     = _normalize(termo).strip()          # minúsculo, sem acento
        termo_digitos = re.sub(r"\D", "", termo)           # só dígitos (CNPJ / pedido)

        # Busca na base COMPLETA (ignora os filtros da sidebar de propósito)
        base = df.copy()

        # Enriquece com vendedor/líder do BKO (cache — não gera chamada extra à API)
        bko_busca = load_bko()
        if not bko_busca.empty and "pedido" in base.columns:
            base = base.merge(
                bko_busca[["pedido", "vendedor_real", "lider"]],
                on="pedido", how="left"
            )

        mask = pd.Series(False, index=base.index)
        if "razao_social" in base.columns and termo_txt:
            mask |= base["razao_social"].apply(lambda x: termo_txt in _normalize(x))
        if termo_digitos:
            if "cnpj" in base.columns:
                mask |= base["cnpj"].apply(lambda x: termo_digitos in re.sub(r"\D", "", _s(x)))
            if "pedido" in base.columns:
                mask |= base["pedido"].apply(lambda x: termo_digitos in re.sub(r"\D", "", _s(x)))

        res = base[mask].copy()

        if res.empty:
            st.warning(f"Nenhum cliente encontrado para «{termo}». A busca cobre toda a base, sem filtros.")
        else:
            ac_res  = int(res["acessos"].sum())      if "acessos"      in res.columns else 0
            rec_res = res["preco_oferta"].sum()       if "preco_oferta" in res.columns else 0
            cli_res = res["razao_social"].nunique()   if "razao_social" in res.columns else 0
            st.caption(
                f"🔎 {len(res)} registro(s) · {cli_res} cliente(s) · "
                f"{ac_res} acessos · R$ {rec_res:,.2f} — busca na base completa (ignora filtros)"
            )

            for c in ["vendedor_real", "lider"]:
                if c in res.columns:
                    res[c] = res[c].apply(lambda x: _s(x) if _s(x) else "—")

            cols_busca = [c for c in [
                "pedido", "razao_social", "cnpj", "parceiro", "tipo_contratacao",
                "vendedor_real", "lider", "mes_ativacao", "acessos", "preco_oferta", "fila_atual"
            ] if c in res.columns]

            ordena = "acessos" if "acessos" in res.columns else cols_busca[0]
            st.dataframe(
                res[cols_busca].sort_values(ordena, ascending=False),
                use_container_width=True, hide_index=True,
                column_config={
                    "pedido":           "Pedido",
                    "razao_social":     "Razão Social",
                    "cnpj":             "CNPJ",
                    "parceiro":         "Parceiro",
                    "tipo_contratacao": "Tipo",
                    "vendedor_real":    "Vendedor",
                    "lider":            "Líder",
                    "mes_ativacao":     "Mês",
                    "acessos":          st.column_config.NumberColumn("Acessos", format="%d"),
                    "preco_oferta":     st.column_config.NumberColumn("Receita", format="R$ %.2f"),
                    "fila_atual":       "Fila Atual",
                },
            )

    # ── Gráfico evolução mensal ───────────────────────────────────────────
    if "mes_ativacao" in df.columns:
        st.markdown('<p class="section-title">📊 Evolução Mensal</p>', unsafe_allow_html=True)

        metrica_graf = st.radio(
            "Exibir no gráfico:",
            ["Acessos", "Receita", "Ambos"],
            horizontal=True,
            label_visibility="collapsed"
        )

        df_graf = df.copy()
        if tipos_sel and "tipo_contratacao" in df_graf.columns:
            df_graf = df_graf[df_graf["tipo_contratacao"].apply(lambda x: _s(x).upper() in tipos_sel)]
        if parceiro_sel != "Todos" and "parceiro" in df_graf.columns:
            df_graf = df_graf[df_graf["parceiro"].apply(lambda x: _s(x) == parceiro_sel)]

        mensal = (df_graf.groupby("mes_ativacao", as_index=False)
                  .agg(Acessos=("acessos","sum"), Receita=("preco_oferta","sum"))
                  .dropna(subset=["mes_ativacao"]))

        try:
            mensal["_dt"] = pd.to_datetime(mensal["mes_ativacao"], format="%m/%Y")
            mensal = mensal.sort_values("_dt").drop(columns=["_dt"])
        except Exception:
            mensal = mensal.sort_values("mes_ativacao")

        if not mensal.empty:
            import altair as alt

            if metrica_graf == "Acessos":
                chart = alt.Chart(mensal).mark_bar(
                    color="#6366f1", cornerRadiusTopLeft=4, cornerRadiusTopRight=4
                ).encode(
                    x=alt.X("mes_ativacao:O", title="Mês", sort=None),
                    y=alt.Y("Acessos:Q", title="Acessos"),
                    tooltip=["mes_ativacao", "Acessos"]
                ).properties(height=280)

            elif metrica_graf == "Receita":
                chart = alt.Chart(mensal).mark_bar(
                    color="#10b981", cornerRadiusTopLeft=4, cornerRadiusTopRight=4
                ).encode(
                    x=alt.X("mes_ativacao:O", title="Mês", sort=None),
                    y=alt.Y("Receita:Q", title="Receita (R$)"),
                    tooltip=["mes_ativacao", "Receita"]
                ).properties(height=280)

            else:  # Ambos
                base = alt.Chart(mensal).encode(x=alt.X("mes_ativacao:O", title="Mês", sort=None))
                bars = base.mark_bar(
                    color="#6366f1", opacity=0.85, cornerRadiusTopLeft=4, cornerRadiusTopRight=4
                ).encode(
                    y=alt.Y("Acessos:Q", title="Acessos", axis=alt.Axis(titleColor="#6366f1")),
                    tooltip=["mes_ativacao","Acessos","Receita"]
                )
                line = base.mark_line(
                    color="#10b981", strokeWidth=3,
                    point=alt.OverlayMarkDef(color="#10b981", size=60)
                ).encode(
                    y=alt.Y("Receita:Q", title="Receita (R$)", axis=alt.Axis(titleColor="#10b981"))
                )
                chart = alt.layer(bars, line).resolve_scale(y="independent").properties(height=280)

            st.altair_chart(
                chart.configure_view(fill="#1a1f2e")
                     .configure_axis(labelColor="#94a3b8", titleColor="#64748b",
                                     gridColor="#2d3748", domainColor="#2d3748"),
                use_container_width=True
            )
        else:
            st.info("Sem dados para exibir no gráfico.")

    # ── Ranking por Vendedor e Equipe ────────────────────────────────────
    st.markdown("---")
    bko      = load_bko()
    colab    = load_colaboradores()
    meta_dict = dict(zip(colab["vendedor"], colab["meta"])) if not colab.empty else {}
    lideres   = sorted(colab["lider"].dropna().unique().tolist()) if not colab.empty else []

    if not bko.empty and "pedido" in dff.columns:
        dff_bko = dff.merge(bko[["pedido","vendedor_real","lider"]], on="pedido", how="left")

        # FIX 1 — Mês: respeita "Todos" em vez de zerar o ranking.
        if mes_sel != "Todos":
            df_atv_rank = dff_bko[dff_bko["mes_ativacao"] == mes_sel].copy()
        else:
            df_atv_rank = dff_bko.copy()

        # FIX 2 — Normaliza lider: vazio/espaços viram ausente (vendas de parceiro).
        if "lider" in df_atv_rank.columns:
            df_atv_rank["lider"] = df_atv_rank["lider"].replace(r"^\s*$", pd.NA, regex=True)

        # FIX 3 — Filtro de líderes: só restringe se o usuário REDUZIU a seleção,
        # e nunca derruba vendas de parceiro (lider ausente).
        if (lideres_sel
                and set(lideres_sel) != set(lideres_disponiveis)
                and "lider" in df_atv_rank.columns):
            df_atv_rank = df_atv_rank[
                df_atv_rank["lider"].isin(lideres_sel) | df_atv_rank["lider"].isna()
            ]

        # Rótulo para as vendas de parceiro aparecerem agrupadas no ranking.
        if "lider" in df_atv_rank.columns:
            df_atv_rank["lider"] = df_atv_rank["lider"].fillna("Parceiros")

        if not df_atv_rank.empty and "vendedor_real" in df_atv_rank.columns:
            # Passa os líderes realmente presentes (inclui "Parceiros") para o ranking por equipe.
            lideres_render = sorted(df_atv_rank["lider"].dropna().unique().tolist())
            render_ranking_resultados(df_atv_rank, mes_sel, meta_dict, lideres_render)
        else:
            st.info("Nenhum registro ativado com BKO para este mês.")
    else:
        st.info("BKO não disponível — cadastre os vendedores na planilha BKO-VENDEDOR-REAL.")

    st.markdown("---")

    # ── Exportar Relatório ────────────────────────────────────────────────
    st.markdown('<p class="section-title">📥 Exportar Relatório de Vendas</p>', unsafe_allow_html=True)

    if not bko.empty and "pedido" in dff.columns and not df_atv_rank.empty:
        # Monta relatório: Equipe → Vendedor → Vendas
        rows_rel = []
        for lider in sorted(df_atv_rank["lider"].dropna().unique()):
            if not lider or lider == "Sem Equipe":
                continue
            df_l = df_atv_rank[df_atv_rank["lider"] == lider]
            for vend in sorted(df_l["vendedor_real"].dropna().unique()):
                if not vend or vend == "Sem Vendedor":
                    continue
                df_v = df_l[df_l["vendedor_real"] == vend]
                for _, row in df_v.iterrows():
                    rows_rel.append({
                        "Equipe / Líder":   lider,
                        "Vendedor":         vend,
                        "Razão Social":     _s(row.get("razao_social", "—")),
                        "Parceiro":         _s(row.get("parceiro", "—")),
                        "Tipo":             _s(row.get("tipo_contratacao", "—")),
                        "Acessos":          int(row.get("acessos", 0)),
                        "Receita (R$)":     float(row.get("preco_oferta", 0)),
                        "Mês Ativação":     _s(row.get("mes_ativacao", "—")),
                        "Fila Atual":       _s(row.get("fila_atual", "—")),
                    })

        if rows_rel:
            df_rel = pd.DataFrame(rows_rel)

            # Resumo por equipe e vendedor
            resumo = (df_rel.groupby(["Equipe / Líder", "Vendedor"], as_index=False)
                      .agg(Total_Acessos=("Acessos", "sum"), Total_Receita=("Receita (R$)", "sum"))
                      .sort_values(["Equipe / Líder", "Total_Receita"], ascending=[True, False]))

            col_prev, col_btn = st.columns([3, 1])
            with col_prev:
                st.dataframe(
                    resumo,
                    use_container_width=True,
                    hide_index=True,
                    column_config={
                        "Equipe / Líder":  "Equipe / Líder",
                        "Vendedor":        "Vendedor",
                        "Total_Acessos":   st.column_config.NumberColumn("Acessos", format="%d"),
                        "Total_Receita":   st.column_config.NumberColumn("Receita (R$)", format="R$ %.2f"),
                    }
                )
            with col_btn:
                st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
                csv_rel = df_rel.to_csv(index=False, sep=";", decimal=",").encode("utf-8-sig")
                st.download_button(
                    label="⬇️ Baixar CSV",
                    data=csv_rel,
                    file_name=f"vendas_{mes_sel.replace('/', '_')}.csv",
                    mime="text/csv",
                    type="primary",
                    use_container_width=True,
                )
                st.caption(f"{len(df_rel)} registro(s) · {mes_sel}")
        else:
            st.info("Nenhum registro para exportar com os filtros atuais.")
    else:
        st.info("Aplique os filtros e aguarde o carregamento do BKO para exportar.")

    st.markdown("---")

    # ── Exportar Relatório (PDF / Excel) ──────────────────────────────────
    st.markdown('<p class="section-title">📥 Exportar Relatório (PDF / Excel)</p>', unsafe_allow_html=True)

    if not bko.empty and "pedido" in dff.columns and not df_atv_rank.empty:
        col_pdf, col_xlsx = st.columns(2)
        with col_pdf:
            if st.button("📄 Gerar PDF", use_container_width=True, type="primary"):
                with st.spinner("Gerando PDF..."):
                    try:
                        pdf_bytes = gerar_pdf_resultados(df_atv_rank, meta_dict, mes_sel)
                        st.download_button(
                            label="⬇️ Baixar PDF",
                            data=pdf_bytes,
                            file_name=f"vendas_{mes_sel.replace('/','_')}.pdf",
                            mime="application/pdf",
                            type="primary",
                            use_container_width=True,
                        )
                    except Exception as e:
                        st.error(f"Erro ao gerar PDF: {e}")
        with col_xlsx:
            if st.button("📊 Gerar Excel", use_container_width=True):
                with st.spinner("Gerando Excel..."):
                    try:
                        xlsx_bytes = gerar_excel_resultados(df_atv_rank, meta_dict, mes_sel)
                        st.download_button(
                            label="⬇️ Baixar Excel",
                            data=xlsx_bytes,
                            file_name=f"vendas_{mes_sel.replace('/','_')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                        )
                    except Exception as e:
                        st.error(f"Erro ao gerar Excel: {e}")
    else:
        st.info("Selecione um mês com dados para exportar.")

    st.markdown("---")
    # ── Tabela detalhada ──────────────────────────────────────────────────
    st.markdown('<p class="section-title">📋 Registros Ativados</p>', unsafe_allow_html=True)

    col_tab1, col_tab2 = st.columns([1, 2])
    with col_tab1:
        st.caption("Por Parceiro")
        if "parceiro" in dff.columns:
            por_parceiro = (dff.groupby("parceiro", as_index=False)
                            .agg(Acessos=("acessos","sum"), Receita=("preco_oferta","sum"))
                            .sort_values("Acessos", ascending=False))
            st.dataframe(por_parceiro, use_container_width=True, hide_index=True,
                column_config={
                    "parceiro": "Parceiro",
                    "Acessos": st.column_config.NumberColumn("Acessos", format="%d"),
                    "Receita": st.column_config.NumberColumn("Receita", format="R$ %.2f"),
                })

    with col_tab2:
        st.caption("Detalhe por Cliente")
        cols_show = [c for c in ["razao_social","parceiro","tipo_contratacao","mes_ativacao","acessos","preco_oferta"]
                     if c in dff.columns]
        st.dataframe(
            dff[cols_show].sort_values("acessos", ascending=False) if "acessos" in dff.columns else dff[cols_show],
            use_container_width=True, hide_index=True,
            column_config={
                "razao_social":     "Razão Social",
                "parceiro":         "Parceiro",
                "tipo_contratacao": "Tipo",
                "mes_ativacao":     "Mês",
                "acessos":          st.column_config.NumberColumn("Acessos", format="%d"),
                "preco_oferta":     st.column_config.NumberColumn("Receita", format="R$ %.2f"),
            })

main()
