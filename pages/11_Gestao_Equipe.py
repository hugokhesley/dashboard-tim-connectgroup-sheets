"""
=====================================================================
  11_Gestao_Equipe.py — Connect Valore · Connect Group
=====================================================================
  Página isolada para líderes e parceiros visualizarem
  apenas suas próprias vendas (Detalhado + Atribuição).

  Níveis de acesso (configurados em secrets [credentials_gestao]):
    - admin   → vê tudo, filtra livremente
    - lider   → vê só a equipe definida em lider = "NOME"
    - parceiro → vê só o parceiro definido em parceiro = "NOME"

  Hierarquia do Detalhado: Líder → Vendedor → Clientes
  (parceiro ignorado na hierarquia — mesmo líder pode vender
   por qualquer parceiro credenciado)
=====================================================================
"""

import streamlit as st
import pandas as pd
import io
from datetime import datetime
import streamlit_authenticator as stauth

from data_loader import (
    load_data, load_bko, load_colaboradores, apply_filters,
    get_parceiros, STATUS_COLORS, _s, _norm_pedido, get_gspread_client
)

import re
import requests

st.set_page_config(
    page_title="Connect Valore — Connect Group",
    page_icon="💎",
    layout="wide",
    initial_sidebar_state="expanded"
)

MESES_PT = {
    "01":"Janeiro","02":"Fevereiro","03":"Março","04":"Abril",
    "05":"Maio","06":"Junho","07":"Julho","08":"Agosto",
    "09":"Setembro","10":"Outubro","11":"Novembro","12":"Dezembro"
}

MES_ALVO          = datetime.now().strftime("%m/%Y")
META_VENDEDOR_PAD = 850
SPREADSHEET_ID    = "1HmtEFf2Akh7NLR2prxDh9S4gmioKYw419B4bkx4yBLg"

# ─────────────────────────────────────────────────────────────────
#  CSS
# ─────────────────────────────────────────────────────────────────

st.markdown("<style>[data-testid='stSidebarNav'] { display: none; }</style>", unsafe_allow_html=True)

st.markdown("""
<style>
  @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');
  html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
  .stApp { background-color: #0f1117; color: #e2e8f0; }
  .header-gestao {
    background: linear-gradient(135deg, #0d2b1a 0%, #14532d 40%, #15803d 70%, #22c55e 100%);
    border-radius: 16px; padding: 24px 32px; margin-bottom: 24px;
    border: 1px solid rgba(255,255,255,0.08);
    display: flex; align-items: center; justify-content: space-between;
  }
  .header-title { font-size: 1.6rem; font-weight: 800; color: #fff; margin: 0; }
  .header-sub   { font-size: 0.82rem; color: rgba(255,255,255,0.65); margin: 4px 0 0; }
  .header-badge { background: rgba(255,255,255,0.15); border: 1px solid rgba(255,255,255,0.25); border-radius: 20px; padding: 5px 14px; font-size: 0.78rem; color: #fff; font-weight: 600; }
  .kpi-mini { background: #1a1f2e; border-radius: 12px; padding: 16px 18px; border: 1px solid #2d3748; position: relative; overflow: hidden; margin-bottom: 4px; }
  .kpi-mini::before { content:''; position:absolute; top:0; left:0; right:0; height:3px; }
  .kpi-mini.blue::before   { background: linear-gradient(90deg, #3b82f6, #1d4ed8); }
  .kpi-mini.green::before  { background: linear-gradient(90deg, #22c55e, #15803d); }
  .kpi-mini.amber::before  { background: linear-gradient(90deg, #f59e0b, #d97706); }
  .kpi-mini.purple::before { background: linear-gradient(90deg, #8b5cf6, #6d28d9); }
  .kpi-label { font-size: 0.68rem; text-transform: uppercase; letter-spacing: 1px; color: #94a3b8; font-weight: 600; margin-bottom: 6px; }
  .kpi-value { font-size: 1.6rem; font-weight: 800; color: #f1f5f9; line-height: 1; }
  .kpi-sub   { font-size: 0.72rem; color: #64748b; margin-top: 4px; }
  .section-title { font-size:0.75rem; text-transform:uppercase; letter-spacing:1.5px; color:#22c55e; font-weight:700; margin:24px 0 12px 0; border-left: 3px solid #22c55e; padding-left: 10px; }
  .det-lider { background: linear-gradient(90deg,#0d1f3c,#1e3a5f); border-left:5px solid #3b82f6; border-radius:12px; padding:14px 22px; margin:20px 0 4px 0; display:flex; align-items:center; justify-content:space-between; }
  .det-lider-nome  { font-size:1.05rem; font-weight:800; color:#93c5fd; }
  .det-lider-stats { font-size:0.78rem; color:#64748b; }
  .det-vendedor-wrap { background:#0f1820; border-left:3px solid #334155; border-radius:8px; padding:10px 16px; margin:4px 0 4px 24px; }
  .det-vend-header { display:flex; align-items:center; justify-content:space-between; margin-bottom:10px; flex-wrap:wrap; gap:6px; }
  .det-vend-nome { font-size:0.83rem; font-weight:700; color:#e2e8f0; }
  .det-vend-kpis { display:flex; gap:0; font-size:0.74rem; background:#0a1018; border-radius:8px; overflow:hidden; border:1px solid #1e293b; }
  .det-kpi-block { display:flex; flex-direction:column; align-items:center; padding:5px 14px; border-right:1px solid #1e293b; line-height:1.3; }
  .det-kpi-block:last-child { border-right:none; }
  .det-kpi-label { font-size:0.62rem; text-transform:uppercase; letter-spacing:.7px; color:#475569; font-weight:600; }
  .det-kpi-val   { font-size:0.82rem; font-weight:700; color:#e2e8f0; margin-top:1px; }
  .det-clientes-wrap { background:#060c12; border-radius:6px; border:1px solid #1a2333; padding:6px 8px; margin-left:4px; }
  .det-cliente-row { display:flex; align-items:center; justify-content:space-between; padding:5px 10px; border-radius:5px; margin:2px 0; background:#0c1520; font-size:0.74rem; border-left:2px solid #1e293b; }
  .det-cli-nome { color:#94a3b8; flex:1; overflow:hidden; text-overflow:ellipsis; white-space:nowrap; margin-right:12px; }
  .det-cli-stats { display:flex; gap:14px; align-items:center; flex-shrink:0; }
  .det-cli-ac   { color:#60a5fa; font-weight:600; font-size:0.73rem; }
  .det-cli-rec  { color:#4ade80; font-weight:600; font-size:0.73rem; }
  .det-cli-fila { font-size:0.66rem; padding:2px 8px; border-radius:99px; font-weight:600; }
  .det-empty { color:#475569; font-size:0.74rem; font-style:italic; padding:4px 6px; }
  .rank-bg   { background: #2d3748; border-radius: 99px; height: 8px; }
  .rank-fill { height: 8px; border-radius: 99px; }
  section[data-testid="stSidebar"] { background: #0d1a0f !important; }
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────
#  AUTENTICAÇÃO
# ─────────────────────────────────────────────────────────────────

def _carregar_auth():
    try:
        creds_raw = dict(st.secrets["credentials_gestao"]["usernames"])
        usernames = {}
        for user, info in creds_raw.items():
            usernames[user] = {
                "name":     info.get("name", user),
                "password": info.get("password", ""),
            }
        return {
            "credentials": {"usernames": usernames},
            "cookie":      {"name": "gestao_cookie", "key": "gestao_secret_key", "expiry_days": 1},
        }
    except Exception as e:
        st.error(f"Erro ao carregar credenciais: {e}")
        st.stop()


def _info_usuario(username: str) -> dict:
    try:
        info = dict(st.secrets["credentials_gestao"]["usernames"][username])
        return {
            "tipo":     info.get("tipo", "lider"),
            "lider":    info.get("lider", ""),
            "parceiro": info.get("parceiro", ""),
            "name":     info.get("name", username),
        }
    except Exception:
        return {"tipo": "lider", "lider": "", "parceiro": "", "name": username}


# ─────────────────────────────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────────────────────────────

def _cor(pct):
    if pct >= 100: return "#22c55e"
    if pct >= 70:  return "#f59e0b"
    return "#ef4444"

def _meta_vend(nome, meta_dict):
    return meta_dict.get(nome, META_VENDEDOR_PAD)

def _fila_badge(fila):
    fila_up = str(fila).strip().upper() if fila else "—"
    cfg  = STATUS_COLORS.get(fila_up, {"border": "#64748b", "icon": "▪️"})
    cor  = cfg.get("border", "#64748b")
    icon = cfg.get("icon", "▪️")
    return (f'<span class="det-cli-fila" '
            f'style="background:{cor}22;color:{cor};border:1px solid {cor}55">'
            f'{icon} {fila_up}</span>')

def gerar_meses_opcoes():
    hoje = datetime.now()
    meses = []
    for i in range(6):
        m = hoje.month - i
        a = hoje.year
        while m <= 0:
            m += 12
            a -= 1
        meses.append(f"{m:02d}/{a}")
    return meses


# ─────────────────────────────────────────────────────────────────
#  EXPORTAR PDF
# ─────────────────────────────────────────────────────────────────

def gerar_pdf(df_atv, meta_dict, mes):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_RIGHT, TA_CENTER

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
        leftMargin=14*mm, rightMargin=14*mm,
        topMargin=14*mm, bottomMargin=14*mm)

    VERDE_ESC = colors.HexColor("#15803d")
    AZUL      = colors.HexColor("#3b82f6")
    AMBER     = colors.HexColor("#f59e0b")
    VERDE     = colors.HexColor("#22c55e")
    CINZA     = colors.HexColor("#64748b")
    CINZA_ESC = colors.HexColor("#1e293b")
    BRANCO    = colors.white
    FL        = colors.HexColor("#131f2e")
    FC        = colors.HexColor("#080e15")
    TEXTO     = colors.HexColor("#e2e8f0")
    TDIM      = colors.HexColor("#94a3b8")

    def s(name, **kw):
        base = {"fontName": "Helvetica", "fontSize": 9, "textColor": TEXTO, "leading": 13}
        base.update(kw)
        return ParagraphStyle(name, **base)

    sTitle = s("T", fontSize=15, fontName="Helvetica-Bold", textColor=BRANCO)
    sRight = s("R", fontSize=7,  textColor=BRANCO, alignment=TA_RIGHT)
    sLider = s("L", fontSize=9,  fontName="Helvetica-Bold", textColor=BRANCO)
    sVend  = s("V", fontSize=8,  fontName="Helvetica-Bold", textColor=TEXTO)
    sCli   = s("C", fontSize=7,  textColor=TDIM)

    story = []
    agora = datetime.now().strftime("%d/%m/%Y %H:%M")

    h = Table([[Paragraph(f"<b>CONNECT GROUP — Connect Valore {mes}</b>", sTitle),
                Paragraph(f"Gerado em {agora}", sRight)]], colWidths=["70%","30%"])
    h.setStyle(TableStyle([
        ("BACKGROUND",(0,0),(-1,-1),VERDE_ESC),
        ("ROWPADDING",(0,0),(-1,-1),10),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE")
    ]))
    story.append(h)
    story.append(Spacer(1, 6*mm))

    lideres = sorted([l for l in df_atv["lider"].dropna().unique() if l and l not in ("Sem Equipe", "")])

    for lider in lideres:
        df_l      = df_atv[df_atv["lider"] == lider]
        df_l_atv  = df_l[df_l["mes_ativacao"] == mes]
        df_l_pip  = df_l[df_l["mes_ativacao"].isna()]
        ac_l      = int(df_l_atv["acessos"].sum())
        rec_l     = df_l_atv["preco_oferta"].sum()
        pip_l     = int(df_l_pip["acessos"].sum())
        vends     = sorted([v for v in df_l["vendedor_real"].dropna().unique() if v and v not in ("Sem Vendedor", "")])
        pip_str_l = f" | ⏳ {pip_l} pip" if pip_l > 0 else ""

        lr = Table([[Paragraph(f"  {lider}", sLider),
                     Paragraph(f"{len(vends)} vend.   {ac_l} ac / R$ {rec_l:,.2f}{pip_str_l}", sRight)]],
                   colWidths=["55%","45%"])
        lr.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,-1),AZUL),
            ("ROWPADDING",(0,0),(-1,-1),8),
            ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
            ("LINEBELOW",(0,0),(-1,0),2,colors.HexColor("#1d4ed8"))
        ]))
        story.append(lr)

        for vend in vends:
            df_v      = df_l[df_l["vendedor_real"] == vend]
            df_v_atv  = df_v[df_v["mes_ativacao"] == mes]
            df_v_pip  = df_v[df_v["mes_ativacao"].isna()]
            ac_v      = int(df_v_atv["acessos"].sum())
            rec_v     = df_v_atv["preco_oferta"].sum()
            ac_pip_v  = int(df_v_pip["acessos"].sum())
            meta_v    = _meta_vend(vend, meta_dict)
            pct_v     = min(int(rec_v / meta_v * 100), 100) if meta_v > 0 else 0
            cor_v     = VERDE if pct_v >= 100 else AMBER if pct_v >= 70 else colors.HexColor("#ef4444")
            pip_str   = f" | ⏳ {ac_pip_v} pip" if ac_pip_v > 0 else ""

            vr = Table([[Paragraph(f"    {vend}", sVend),
                         Paragraph(f"{ac_v} ac / R$ {rec_v:,.2f} ({pct_v}%){pip_str}", sRight)]],
                       colWidths=["55%","45%"])
            vr.setStyle(TableStyle([
                ("BACKGROUND",(0,0),(-1,-1),FL),
                ("ROWPADDING",(0,0),(-1,-1),6),
                ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
                ("LINEBELOW",(0,0),(-1,0),0.5,CINZA_ESC),
                ("TEXTCOLOR",(1,0),(1,0),cor_v)
            ]))
            story.append(vr)

            cols_grp = [c for c in ["razao_social","fila_atual"] if c in df_v.columns]
            if cols_grp:
                df_g = df_v.copy()
                for col in cols_grp:
                    df_g[col] = df_g[col].fillna("—")
                cdf = (df_g.groupby(cols_grp, as_index=False)
                       .agg(ac=("acessos","sum"), rec=("preco_oferta","sum"))
                       .sort_values("ac", ascending=False))
                for _, crow in cdf.iterrows():
                    razao = str(crow.get("razao_social","—"))[:55]
                    fila  = str(crow.get("fila_atual","—")).upper()
                    ac_c  = int(crow.get("ac",0))
                    rec_c = float(crow.get("rec",0))
                    cr = Table([[Paragraph(f"        {razao}", sCli),
                                 Paragraph(f"{ac_c} ac   R$ {rec_c:,.2f}   {fila}", sRight)]],
                               colWidths=["60%","40%"])
                    cr.setStyle(TableStyle([
                        ("BACKGROUND",(0,0),(-1,-1),FC),
                        ("ROWPADDING",(0,0),(-1,-1),3),
                        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
                        ("LINEBELOW",(0,0),(-1,0),0.3,CINZA_ESC)
                    ]))
                    story.append(cr)

        story.append(Spacer(1, 4*mm))

    story.append(HRFlowable(width="100%", thickness=0.5, color=CINZA_ESC))
    story.append(Spacer(1, 1*mm))
    story.append(Paragraph(
        f"Connect Group · {mes} · {agora}",
        s("F", fontSize=6, textColor=CINZA, alignment=TA_CENTER)
    ))
    doc.build(story)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────────────────────────
#  EXPORTAR EXCEL
# ─────────────────────────────────────────────────────────────────

def gerar_excel(df_atv, meta_dict, mes):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = f"Vendas {mes.replace('/', '-')}"

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color.replace("#",""))

    ft_title = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    ft_lider = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    ft_vend  = Font(name="Calibri", bold=True, size=10, color="E2E8F0")
    ft_cli   = Font(name="Calibri", size=9,    color="94A3B8")

    al_left  = Alignment(horizontal="left",  vertical="center")
    al_right = Alignment(horizontal="right", vertical="center")
    al_center= Alignment(horizontal="center",vertical="center")

    agora = datetime.now().strftime("%d/%m/%Y %H:%M")

    # A=Nome, B=Fila, C=Acessos (total), D=Receita (total)
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 18

    row = 1
    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = f"CONNECT GROUP — Connect Valore {mes}"
    ws[f"A{row}"].font = ft_title
    ws[f"A{row}"].fill = fill("#15803d")
    ws[f"A{row}"].alignment = al_left
    ws.row_dimensions[row].height = 28
    row += 1

    ws.merge_cells(f"A{row}:D{row}")
    ws[f"A{row}"] = f"Gerado em {agora}"
    ws[f"A{row}"].font = Font(name="Calibri", size=9, color="94A3B8")
    ws[f"A{row}"].fill = fill("#0f1117")
    ws[f"A{row}"].alignment = al_left
    row += 1

    row += 1  # linha em branco

    lideres = sorted([l for l in df_atv["lider"].dropna().unique() if l and l not in ("Sem Equipe", "")])

    for lider in lideres:
        df_l = df_atv[df_atv["lider"] == lider]
        # Total = ativados + tramitando
        ac_l  = int(df_l["acessos"].sum())
        rec_l = df_l["preco_oferta"].sum()
        vends = sorted([v for v in df_l["vendedor_real"].dropna().unique() if v and v not in ("Sem Vendedor", "")])

        ws.merge_cells(f"A{row}:B{row}")
        ws[f"A{row}"] = f"  {lider}"
        ws[f"A{row}"].font = ft_lider
        ws[f"A{row}"].fill = fill("#1d4ed8")
        ws[f"A{row}"].alignment = al_left
        ws[f"C{row}"] = ac_l
        ws[f"C{row}"].font = ft_lider
        ws[f"C{row}"].fill = fill("#1d4ed8")
        ws[f"C{row}"].alignment = al_right
        ws[f"D{row}"] = rec_l
        ws[f"D{row}"].font = ft_lider
        ws[f"D{row}"].fill = fill("#1d4ed8")
        ws[f"D{row}"].alignment = al_right
        ws[f"D{row}"].number_format = "R$ #,##0.00"
        ws.row_dimensions[row].height = 20
        row += 1

        for vend in vends:
            df_v   = df_l[df_l["vendedor_real"] == vend]
            # Total = ativados + tramitando
            ac_v   = int(df_v["acessos"].sum())
            rec_v  = df_v["preco_oferta"].sum()
            # % meta usa só ativados do mes
            df_v_atv = df_v[df_v["mes_ativacao"] == mes]
            rec_v_atv = df_v_atv["preco_oferta"].sum()
            meta_v = _meta_vend(vend, meta_dict)
            pct_v  = min(int(rec_v_atv / meta_v * 100), 100) if meta_v > 0 else 0
            cor_v  = "22C55E" if pct_v >= 100 else "F59E0B" if pct_v >= 70 else "EF4444"

            ws.merge_cells(f"A{row}:B{row}")
            ws[f"A{row}"] = f"    {vend}"
            ws[f"A{row}"].font = ft_vend
            ws[f"A{row}"].fill = fill("#131f2e")
            ws[f"A{row}"].alignment = al_left
            ws[f"C{row}"] = ac_v
            ws[f"C{row}"].font = Font(name="Calibri", bold=True, size=10, color=cor_v)
            ws[f"C{row}"].fill = fill("#131f2e")
            ws[f"C{row}"].alignment = al_right
            ws[f"D{row}"] = rec_v
            ws[f"D{row}"].font = Font(name="Calibri", bold=True, size=10, color=cor_v)
            ws[f"D{row}"].fill = fill("#131f2e")
            ws[f"D{row}"].alignment = al_right
            ws[f"D{row}"].number_format = "R$ #,##0.00"
            ws.row_dimensions[row].height = 18
            row += 1

            # Clientes — total por cliente (ativados + tramitando)
            cols_grp = [c for c in ["razao_social", "fila_atual"] if c in df_v.columns]
            if cols_grp:
                df_g = df_v.copy()
                for col in cols_grp:
                    df_g[col] = df_g[col].fillna("—")
                cdf = (df_g.groupby(cols_grp, as_index=False)
                       .agg(ac=("acessos","sum"), rec=("preco_oferta","sum"))
                       .sort_values("ac", ascending=False))
                for _, crow in cdf.iterrows():
                    razao = str(crow.get("razao_social","—"))[:60]
                    fila  = str(crow.get("fila_atual","—")).upper()
                    ac_c  = int(crow.get("ac",0))
                    rec_c = float(crow.get("rec",0))

                    ws[f"A{row}"] = f"        {razao}"
                    ws[f"A{row}"].font = ft_cli
                    ws[f"A{row}"].fill = fill("#080e15")
                    ws[f"A{row}"].alignment = al_left
                    ws[f"B{row}"] = fila
                    ws[f"B{row}"].font = ft_cli
                    ws[f"B{row}"].fill = fill("#080e15")
                    ws[f"B{row}"].alignment = al_center
                    ws[f"C{row}"] = ac_c
                    ws[f"C{row}"].font = ft_cli
                    ws[f"C{row}"].fill = fill("#080e15")
                    ws[f"C{row}"].alignment = al_right
                    ws[f"D{row}"] = rec_c
                    ws[f"D{row}"].font = ft_cli
                    ws[f"D{row}"].fill = fill("#080e15")
                    ws[f"D{row}"].alignment = al_right
                    ws[f"D{row}"].number_format = "R$ #,##0.00"
                    ws.row_dimensions[row].height = 15
                    row += 1

        row += 1  # linha em branco entre lideres

    buf = __import__('io').BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────────────────────────
#  RENDER DETALHADO — hierarquia: Líder → Vendedor → Clientes
#  (parceiro ignorado — mesmo líder pode vender por qualquer parceiro)
# ─────────────────────────────────────────────────────────────────

def render_detalhado(df, mes_alvo, meta_dict):
    FILAS_CANCEL = {"CANCELADO", "CANCELADA", "CANCEL"}

    if "fila_atual" in df.columns:
        df = df[~df["fila_atual"].str.strip().str.upper().isin(FILAS_CANCEL)]

    lideres = sorted([l for l in df["lider"].dropna().unique() if l and l not in ("Sem Equipe", "")])

    if not lideres:
        st.info("Nenhuma equipe encontrada para os filtros selecionados.")
        return

    for lider in lideres:
        df_l    = df[df["lider"] == lider]
        ac_l    = int(df_l[df_l["mes_ativacao"] == mes_alvo]["acessos"].sum())
        rec_l   = df_l[df_l["mes_ativacao"] == mes_alvo]["preco_oferta"].sum()
        pip_l   = int(df_l[df_l["mes_ativacao"].isna()]["acessos"].sum())
        vends   = sorted([v for v in df_l["vendedor_real"].dropna().unique() if v and v not in ("Sem Vendedor", "")])

        st.markdown(f"""<div class="det-lider">
          <span class="det-lider-nome">👤 {lider}</span>
          <span class="det-lider-stats">
            {len(vends)} vend. &nbsp;·&nbsp;
            ✅ {ac_l} ac / R$ {rec_l:,.2f} &nbsp;·&nbsp;
            ⏳ {pip_l} ac tramitando
          </span>
        </div>""", unsafe_allow_html=True)

        for vend in vends:
            df_v      = df_l[df_l["vendedor_real"] == vend]
            df_atv_v  = df_v[df_v["mes_ativacao"] == mes_alvo]
            df_pip_v  = df_v[df_v["mes_ativacao"].isna()]
            ac_atv_v  = int(df_atv_v["acessos"].sum())
            rec_atv_v = df_atv_v["preco_oferta"].sum()
            ac_pip_v  = int(df_pip_v["acessos"].sum())
            rec_pip_v = df_pip_v["preco_oferta"].sum()
            meta_v    = _meta_vend(vend, meta_dict)
            pct_v     = min(int(rec_atv_v / meta_v * 100), 100) if meta_v > 0 else 0
            cor_v     = _cor(pct_v)

            cols_grp = [c for c in ["razao_social", "fila_atual", "status_dash"] if c in df_v.columns]
            if cols_grp:
                df_grp = df_v.copy()
                for col in cols_grp:
                    df_grp[col] = df_grp[col].fillna("—")
                clientes_df = (
                    df_grp.groupby(cols_grp, as_index=False)
                    .agg(ac=("acessos","sum"), rec=("preco_oferta","sum"),
                         n_atv=("mes_ativacao", lambda x: (x == mes_alvo).sum()))
                    .sort_values(["n_atv","ac"], ascending=[False, False])
                )
            else:
                clientes_df = pd.DataFrame()

            clientes_html = ""
            if clientes_df.empty:
                clientes_html = '<div class="det-empty">Sem pedidos no mês.</div>'
            else:
                for _, crow in clientes_df.iterrows():
                    razao = str(crow.get("razao_social", "—"))
                    if razao.strip().startswith("#") or razao.strip() == "":
                        razao = "—"
                    fila        = str(crow.get("fila_atual", crow.get("status_dash", "—")))
                    ac_c        = int(crow.get("ac", 0))
                    rec_c       = float(crow.get("rec", 0))
                    razao_trunc = (razao[:52] + "…") if len(razao) > 52 else razao
                    clientes_html += f"""<div class="det-cliente-row">
                      <span class="det-cli-nome">{razao_trunc}</span>
                      <span class="det-cli-stats">
                        <span class="det-cli-ac">🔢 {ac_c} ac</span>
                        <span class="det-cli-rec">💰 R$ {rec_c:,.2f}</span>
                        {_fila_badge(fila)}
                      </span>
                    </div>"""

            st.markdown(f"""<div class="det-vendedor-wrap">
              <div class="det-vend-header">
                <span class="det-vend-nome">📌 {vend}</span>
                <div class="det-vend-kpis">
                  <div class="det-kpi-block">
                    <span class="det-kpi-label">✅ Ativado</span>
                    <span class="det-kpi-val" style="color:{cor_v}">
                      {ac_atv_v} ac · R$ {rec_atv_v:,.2f}
                      <span style="color:#64748b;font-size:0.68rem"> ({pct_v}%)</span>
                    </span>
                  </div>
                  <div class="det-kpi-block">
                    <span class="det-kpi-label">⏳ Tramitando</span>
                    <span class="det-kpi-val" style="color:#f59e0b">
                      {ac_pip_v} ac · R$ {rec_pip_v:,.2f}
                    </span>
                  </div>
                </div>
              </div>
              <div class="det-clientes-wrap">{clientes_html}</div>
            </div>""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────
#  RENDER ATRIBUIÇÃO
# ─────────────────────────────────────────────────────────────────

def render_atribuicao(df_pendentes, ws_bko, vendedores):
    if df_pendentes.empty:
        st.success("✅ Nenhum pedido pendente de atribuição!")
        return

    total = len(df_pendentes)
    st.markdown(f"**{total} pedido(s) sem vendedor.** Preencha os do seu time e salve.")
    st.markdown("---")

    vendedores_opts = ["— não é meu —"] + vendedores
    atribuicoes = []

    for _, row in df_pendentes.iterrows():
        pedido  = _s(row.get("pedido", ""))
        cliente = _s(row.get("razao_social", "—"))
        safra   = _s(row.get("safra", "—"))
        row_idx = int(row.get("_row_idx", 0))
        col_idx = int(row.get("_col_vendedor_idx", 4))

        col1, col2 = st.columns([3, 2])
        with col1:
            st.markdown(f"""<div style="background:#1a1f2e;border:1px solid #2d3748;border-radius:12px;padding:14px 16px;margin-bottom:4px">
              <div style="font-size:0.75rem;color:#64748b">Pedido {pedido} · {safra}</div>
              <div style="font-size:0.95rem;font-weight:600;color:#f1f5f9;margin-top:4px">{cliente}</div>
            </div>""", unsafe_allow_html=True)
        with col2:
            sel = st.selectbox(" ", vendedores_opts, key=f"atr_{pedido}", label_visibility="collapsed")
            if sel and sel != "— não é meu —":
                atribuicoes.append({"pedido": pedido, "row_idx": row_idx, "col_idx": col_idx, "vendedor": sel})

    st.markdown("---")
    n_attr = len(atribuicoes)
    if n_attr > 0:
        st.info(f"**{n_attr}** pedido(s) atribuído(s). Salve quando terminar.")
        if st.button("💾 Salvar atribuições", type="primary", use_container_width=True):
            salvos = 0
            with st.spinner("Salvando..."):
                for item in atribuicoes:
                    try:
                        ws_bko.update_cell(item["row_idx"], item["col_idx"], item["vendedor"])
                        salvos += 1
                    except Exception as e:
                        st.warning(f"Erro no pedido {item['pedido']}: {e}")
                st.cache_data.clear()
            st.success(f"✅ {salvos} atribuição(ões) salva(s)!")
            st.rerun()
    else:
        st.button("💾 Salvar atribuições", disabled=True, use_container_width=True)
        st.caption("Selecione ao menos um vendedor para salvar.")

    if st.button("🔄 Atualizar lista", use_container_width=True):
        st.cache_data.clear()
        st.rerun()


# ─────────────────────────────────────────────────────────────────
#  EXPORTAR — botões PDF e Excel
# ─────────────────────────────────────────────────────────────────

def _render_exportar(df_export, meta_dict, mes_alvo):
    if df_export.empty or "lider" not in df_export.columns:
        st.info("Sem dados ativados no mês para exportar.")
        return

    col_pdf, col_xlsx = st.columns(2)

    with col_pdf:
        if st.button("📄 Gerar PDF", use_container_width=True, type="primary"):
            with st.spinner("Gerando PDF..."):
                try:
                    pdf_bytes = gerar_pdf(df_export, meta_dict, mes_alvo)
                    st.download_button(
                        label="⬇️ Baixar PDF",
                        data=pdf_bytes,
                        file_name=f"gestao_vendas_{mes_alvo.replace('/','_')}.pdf",
                        mime="application/pdf",
                        type="primary",
                        use_container_width=True,
                    )
                except Exception as e:
                    st.error(f"Erro ao gerar PDF: {e}")

    with col_xlsx:
        if st.button("📊 Gerar Excel", use_container_width=True):
            with st.spinner("Gerando Excel..."):
                try:
                    xlsx_bytes = gerar_excel(df_export, meta_dict, mes_alvo)
                    st.download_button(
                        label="⬇️ Baixar Excel",
                        data=xlsx_bytes,
                        file_name=f"gestao_vendas_{mes_alvo.replace('/','_')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
                except Exception as e:
                    st.error(f"Erro ao gerar Excel: {e}")


# ─────────────────────────────────────────────────────────────────
#  CARTEIRA DE ATENDIMENTO — constantes, helpers e UI
# ─────────────────────────────────────────────────────────────────

ABA_CARTEIRA    = "CarteiraAtendimento"
DIAS_EXPIRACAO  = 60
STATUS_CARTEIRA = ["Em Atendimento", "Ativado", "Expirado", "Liberado pelo Admin"]
COR_STATUS_CART = {
    "Em Atendimento":      "#f59e0b",
    "Ativado":             "#22c55e",
    "Expirado":            "#64748b",
    "Liberado pelo Admin": "#8b5cf6",
}
HEADER_CARTEIRA = [
    "cnpj","razao_social","nome_fantasia","atividade_economica",
    "data_fundacao","capital_social","telefone","email",
    "cep","logradouro","numero","bairro","cidade","uf",
    "vendedor","lider","tbp",
    "status","data_registro","data_atualizacao",
    "pedido_tim","obs","registrado_por",
]


def _formatar_cnpj(cnpj: str) -> str:
    c = re.sub(r"\D","",str(cnpj))
    if len(c) == 14:
        return f"{c[:2]}.{c[2:5]}.{c[5:8]}/{c[8:12]}-{c[12:]}"
    return cnpj


def _dias_restantes_cart(data_registro: str) -> int:
    try:
        dt = datetime.strptime(str(data_registro).strip(), "%d/%m/%Y %H:%M")
        return max(DIAS_EXPIRACAO - (datetime.now() - dt).days, 0)
    except Exception:
        return DIAS_EXPIRACAO


@st.cache_data(ttl=30, show_spinner=False)
def _load_carteira(_gc):
    try:
        sh = _gc.open_by_key(SPREADSHEET_ID)
        try:
            aba = sh.worksheet(ABA_CARTEIRA)
        except Exception:
            aba = sh.add_worksheet(title=ABA_CARTEIRA, rows=2000, cols=len(HEADER_CARTEIRA)+2)
            aba.append_row(HEADER_CARTEIRA)
            return pd.DataFrame(columns=HEADER_CARTEIRA)
        rows = aba.get_all_records()
        if not rows:
            return pd.DataFrame(columns=HEADER_CARTEIRA)
        df = pd.DataFrame(rows)
        for col in HEADER_CARTEIRA:
            if col not in df.columns:
                df[col] = ""
        return df
    except Exception as e:
        st.error(f"Erro ao carregar carteira: {e}")
        return pd.DataFrame(columns=HEADER_CARTEIRA)


@st.cache_data(ttl=120, show_spinner=False)
def _load_colab_carteira():
    """Carrega vendedores com lider ativo da aba Colaboradores."""
    import unicodedata
    def _n(s):
        s = str(s).strip().lower()
        return "".join(c for c in unicodedata.normalize("NFD",s) if unicodedata.category(c)!="Mn")
    try:
        df = load_colaboradores()
        if df.empty:
            return [], {}, {}
        col_map = {_n(c): c for c in df.columns}
        col_v = col_map.get("vendedor")
        col_l = col_map.get("lider")
        col_t = col_map.get("tbp")
        if not col_v:
            return [], {}, {}
        vendedores, mapa_lider, mapa_tbp = [], {}, {}
        for _, row in df.iterrows():
            v = str(row.get(col_v,"")).strip()
            l = str(row.get(col_l,"")).strip() if col_l else ""
            t = str(row.get(col_t,"")).strip() if col_t else ""
            # Só inclui vendedor que tem líder definido (ativo)
            if v and v not in ("","nan","None","VENDEDOR") and l and l not in ("","nan","None","LIDER","LÍDER"):
                mapa_lider[v] = l
                if t and t not in ("","nan","None","TBP"):
                    mapa_tbp[v] = t
        vendedores = sorted(mapa_lider.keys())
        return vendedores, mapa_lider, mapa_tbp
    except Exception:
        return [], {}, {}


@st.cache_data(ttl=3600, show_spinner=False)
def _buscar_cnpj_receita(cnpj: str) -> dict:
    cnpj_limpo = re.sub(r"\D","",cnpj)
    if len(cnpj_limpo) != 14:
        return {"erro": "CNPJ deve ter 14 digitos"}
    try:
        r = requests.get(f"https://brasilapi.com.br/api/cnpj/v1/{cnpj_limpo}", timeout=10)
        if r.status_code == 200:
            d = r.json()
            return {
                "cnpj":                cnpj_limpo,
                "razao_social":        d.get("razao_social",""),
                "nome_fantasia":       d.get("nome_fantasia","") or d.get("razao_social",""),
                "atividade_economica": (d.get("cnae_fiscal_descricao") or "")[:60],
                "data_fundacao":       d.get("data_inicio_atividade",""),
                "capital_social":      f"R$ {d.get('capital_social',0):,.2f}",
                "telefone":            d.get("ddd_telefone_1",""),
                "email":               d.get("email",""),
                "cep":                 re.sub(r"\D","",d.get("cep","")),
                "logradouro":          d.get("logradouro",""),
                "numero":              d.get("numero",""),
                "bairro":              d.get("bairro",""),
                "cidade":              d.get("municipio",""),
                "uf":                  d.get("uf",""),
            }
        return {"erro": "CNPJ nao encontrado na Receita Federal"}
    except Exception as e:
        return {"erro": str(e)}


def _cnpj_disponivel(df: pd.DataFrame, cnpj: str):
    cnpj_limpo = re.sub(r"\D","",cnpj)
    if df.empty:
        return True, None
    match = df[df["cnpj"].astype(str).str.replace(r"\D","",regex=True) == cnpj_limpo]
    if match.empty:
        return True, None
    row = match.iloc[0]
    if str(row.get("status","")) == "Em Atendimento":
        return False, row.to_dict()
    return True, row.to_dict()


def _atualizar_expirados_gc(gc, df: pd.DataFrame) -> pd.DataFrame:
    hoje = datetime.now()
    atualizados = []
    for idx, row in df.iterrows():
        if str(row.get("status","")) != "Em Atendimento":
            continue
        try:
            dt = datetime.strptime(str(row.get("data_registro","")).strip(), "%d/%m/%Y %H:%M")
            if (hoje - dt).days >= DIAS_EXPIRACAO:
                df.at[idx,"status"] = "Expirado"
                df.at[idx,"data_atualizacao"] = hoje.strftime("%d/%m/%Y %H:%M")
                atualizados.append(str(row.get("cnpj","")))
        except Exception:
            pass
    if atualizados:
        try:
            sh    = gc.open_by_key(SPREADSHEET_ID)
            aba   = sh.worksheet(ABA_CARTEIRA)
            todos = aba.get_all_values()
            header= [str(h).strip() for h in todos[0]] if todos else []
            if "status" in header and "cnpj" in header:
                col_st = header.index("status")+1
                col_da = header.index("data_atualizacao")+1
                col_cn = header.index("cnpj")+1
                for i, r2 in enumerate(todos[1:], start=2):
                    if re.sub(r"\D","",str(r2[col_cn-1]).strip()) in [re.sub(r"\D","",c) for c in atualizados]:
                        aba.update_cell(i, col_st, "Expirado")
                        aba.update_cell(i, col_da, hoje.strftime("%d/%m/%Y %H:%M"))
        except Exception:
            pass
    return df


def _registrar_cnpj_carteira(gc, dados: dict, usuario: str) -> bool:
    try:
        sh    = gc.open_by_key(SPREADSHEET_ID)
        aba   = sh.worksheet(ABA_CARTEIRA)
        agora = datetime.now().strftime("%d/%m/%Y %H:%M")
        linha = []
        for c in HEADER_CARTEIRA:
            if c == "status":                          linha.append("Em Atendimento")
            elif c in ("data_registro","data_atualizacao"): linha.append(agora)
            elif c == "registrado_por":                linha.append(usuario)
            else:                                      linha.append(dados.get(c,""))
        aba.append_row(linha)
        st.cache_data.clear()
        return True
    except Exception as e:
        st.error(f"Erro ao registrar: {e}")
        return False


def _atualizar_campo_carteira(gc, cnpj: str, campos: dict) -> bool:
    try:
        sh    = gc.open_by_key(SPREADSHEET_ID)
        aba   = sh.worksheet(ABA_CARTEIRA)
        todos = aba.get_all_values()
        if not todos:
            return False
        header     = [str(h).strip() for h in todos[0]]
        cnpj_limpo = re.sub(r"\D","",cnpj)
        col_cn     = header.index("cnpj") if "cnpj" in header else None
        if col_cn is None:
            return False
        campos["data_atualizacao"] = datetime.now().strftime("%d/%m/%Y %H:%M")
        for i, r2 in enumerate(todos[1:], start=2):
            if not r2:
                continue
            if re.sub(r"\D","",str(r2[col_cn]).strip()) == cnpj_limpo:
                for campo, valor in campos.items():
                    if campo in header:
                        aba.update_cell(i, header.index(campo)+1, valor)
                st.cache_data.clear()
                return True
        return False
    except Exception as e:
        st.error(f"Erro ao atualizar: {e}")
        return False


def _card_cliente_carteira(row: dict, is_admin: bool, gc, username: str):
    cnpj    = _formatar_cnpj(str(row.get("cnpj","")))
    razao   = str(row.get("razao_social","—"))
    fantasia= str(row.get("nome_fantasia",""))
    vend    = str(row.get("vendedor","—"))
    lider_c = str(row.get("lider","—"))
    tbp     = str(row.get("tbp","—"))
    status  = str(row.get("status","Em Atendimento"))
    data_r  = str(row.get("data_registro",""))
    pedido  = str(row.get("pedido_tim",""))
    obs     = str(row.get("obs",""))
    cidade  = str(row.get("cidade",""))
    uf      = str(row.get("uf",""))
    cor     = COR_STATUS_CART.get(status,"#64748b")
    dias    = _dias_restantes_cart(data_r) if status == "Em Atendimento" else None

    dias_html = pedido_html = obs_html = ""
    if dias is not None:
        cor_d = "#22c55e" if dias>30 else "#f59e0b" if dias>10 else "#ef4444"
        dias_html = f"<span style='color:{cor_d};font-weight:700;font-size:0.73rem'>⏱ {dias} dias restantes</span>"
    if pedido and pedido not in ("","nan"):
        pedido_html = f"<span style='color:#60a5fa;font-size:0.73rem'>📌 TIM: {pedido}</span>"
    if obs and obs not in ("","nan"):
        obs_html = f"<div style='font-size:0.72rem;color:#475569;margin-top:4px'>📝 {obs}</div>"
    fantasia_html = f"{fantasia} · " if fantasia and fantasia != razao else ""

    st.markdown(f"""
    <div style="background:#111827;border:1px solid #1e3a5f;border-left:5px solid {cor};
                border-radius:12px;padding:14px 18px;margin-bottom:8px">
      <div style="display:flex;justify-content:space-between;align-items:start;flex-wrap:wrap;gap:8px">
        <div style="flex:1;min-width:200px">
          <div style="font-size:0.95rem;font-weight:700;color:#f1f5f9">{razao}</div>
          <div style="font-size:0.73rem;color:#64748b;margin-top:2px">{fantasia_html}{cnpj} · {cidade}/{uf}</div>
          <div style="font-size:0.73rem;color:#64748b;margin-top:5px">
            👤 <b style="color:#e2e8f0">{vend}</b> · 🏅 {lider_c} · 🏢 <span style="color:#a78bfa">{tbp}</span>
          </div>
          <div style="margin-top:6px;display:flex;gap:12px;align-items:center;flex-wrap:wrap">
            {dias_html} {pedido_html}
          </div>
          {obs_html}
        </div>
        <div style="display:flex;flex-direction:column;align-items:flex-end;gap:6px">
          <span style="display:inline-block;padding:3px 10px;border-radius:99px;
                       font-size:0.7rem;font-weight:700;color:#fff;background:{cor}">{status}</span>
          <span style="font-size:0.67rem;color:#475569">{data_r[:10] if data_r else ""}</span>
        </div>
      </div>
    </div>
    """, unsafe_allow_html=True)


    cnpj_key = re.sub(r"\D", "", str(row.get("cnpj","")))
    key_dossie = f"dossie_{cnpj_key}"

    # Botões de ação
    col_atend, col_edit = st.columns([1, 1])
    with col_atend:
        if st.button("📋 Atendimento", key=f"btn_atend_{cnpj_key}",
                     use_container_width=True, type="primary"):
            if st.session_state.get(key_dossie):
                del st.session_state[key_dossie]
            else:
                st.session_state[key_dossie] = True
            st.rerun()
    with col_edit:
        if st.button("⚙️ Editar", key=f"btn_edit_{cnpj_key}", use_container_width=True):
            toggle_key = f"edit_open_{cnpj_key}"
            st.session_state[toggle_key] = not st.session_state.get(toggle_key, False)
            st.rerun()

    # Painel de edição admin
    if st.session_state.get(f"edit_open_{cnpj_key}"):
        c1, c2, c3, c4 = st.columns([2,2,3,1])
        with c1:
            ns = st.selectbox("Status", STATUS_CARTEIRA,
                index=STATUS_CARTEIRA.index(status) if status in STATUS_CARTEIRA else 0,
                key=f"cst_{cnpj_key}")
        with c2:
            np_ = st.text_input("Nº Pedido TIM",
                value=pedido if pedido not in ("","nan") else "",
                key=f"cpt_{cnpj_key}")
        with c3:
            no_ = st.text_input("Observação",
                value=obs if obs not in ("","nan") else "",
                key=f"cob_{cnpj_key}")
        with c4:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("💾", key=f"csv_{cnpj_key}", use_container_width=True):
                ok = _atualizar_campo_carteira(gc, cnpj_key,
                                               {"status":ns,"pedido_tim":np_,"obs":no_})
                if ok:
                    st.success("✅ Salvo!")
                    st.session_state.pop(f"edit_open_{cnpj_key}", None)
                    st.rerun()
    elif status == "Em Atendimento" and pedido in ("","nan"):
        col_p1, col_p2 = st.columns([3,1])
        with col_p1:
            np2 = st.text_input("Nº Pedido TIM *", placeholder="ex: 6341069",
                                key=f"cnp2_{cnpj_key}")
        with col_p2:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("📌 Registrar", key=f"cnpbt_{cnpj_key}",
                         type="primary", use_container_width=True):
                if np2.strip():
                    ok = _atualizar_campo_carteira(gc, cnpj_key,
                                                   {"pedido_tim": np2.strip()})
                    if ok:
                        st.success("📌 Registrado!")
                        st.rerun()
                else:
                    st.error("Informe o número.")

    # Dossiê expandido
    if st.session_state.get(key_dossie):
        st.markdown("---")
        if st.button("✖ Fechar dossiê", key=f"fechar_{cnpj_key}"):
            del st.session_state[key_dossie]
            st.rerun()
        _tela_dossie(gc, cnpj_key, row, username, is_admin)
        st.markdown("---")



# ─────────────────────────────────────────────────────────────────
#  DOSSIÊ DO CLIENTE — Contatos, Pedidos, Qualidade
# ─────────────────────────────────────────────────────────────────

ABA_CONTATOS = "CarteiraContatos"
HEADER_CONTATOS = [
    "cnpj", "data", "tipo", "responsavel", "obs", "registrado_em"
]
TIPOS_CONTATO = ["📞 Ligação", "🏠 Visita", "💬 WhatsApp",
                  "📧 Email", "📋 Reunião", "📝 Outro"]


def _get_aba_contatos(gc):
    sh = gc.open_by_key(SPREADSHEET_ID)
    try:
        return sh.worksheet(ABA_CONTATOS)
    except Exception:
        aba = sh.add_worksheet(title=ABA_CONTATOS, rows=5000, cols=len(HEADER_CONTATOS)+2)
        aba.append_row(HEADER_CONTATOS)
        return aba


@st.cache_data(ttl=20, show_spinner=False)
def _load_contatos_cnpj(_gc, cnpj_norm: str) -> pd.DataFrame:
    try:
        aba  = _get_aba_contatos(_gc)
        rows = aba.get_all_records()
        if not rows:
            return pd.DataFrame(columns=HEADER_CONTATOS)
        df = pd.DataFrame(rows)
        for col in HEADER_CONTATOS:
            if col not in df.columns:
                df[col] = ""
        df["cnpj"] = df["cnpj"].astype(str).str.replace(r"\D","",regex=True).str.strip()
        return df[df["cnpj"] == cnpj_norm].sort_values(
            "registrado_em", ascending=False).reset_index(drop=True)
    except Exception:
        return pd.DataFrame(columns=HEADER_CONTATOS)


def _salvar_contato(gc, cnpj_norm: str, tipo: str, obs: str, usuario: str) -> bool:
    try:
        aba   = _get_aba_contatos(gc)
        agora = datetime.now().strftime("%d/%m/%Y %H:%M")
        aba.append_row([
            cnpj_norm,
            agora[:10],
            tipo,
            usuario,
            obs,
            agora,
        ])
        st.cache_data.clear()
        return True
    except Exception as e:
        st.error(f"Erro ao salvar contato: {e}")
        return False


@st.cache_data(ttl=60, show_spinner=False)
def _pedidos_por_cnpj(_gc, cnpj_norm: str) -> pd.DataFrame:
    """Busca pedidos do cliente no DadosRadar pelo CNPJ."""
    try:
        sh     = _gc.open_by_key(SPREADSHEET_ID)
        radar  = sh.worksheet("DadosRadar")
        rows   = radar.get_all_values()
        if not rows or len(rows) < 2:
            return pd.DataFrame()
        header = [str(h).strip().lower() for h in rows[0]]
        n = len(header)
        data = [r + [""]*(n-len(r)) if len(r)<n else r[:n] for r in rows[1:]]
        df = pd.DataFrame(data, columns=header)
        ci = next((i for i,h in enumerate(header) if h=="cnpj"), None)
        if ci is None:
            return pd.DataFrame()
        df["_cnpj_norm"] = df.iloc[:,ci].astype(str).str.replace(r"\D","",regex=True)
        return df[df["_cnpj_norm"] == cnpj_norm].reset_index(drop=True)
    except Exception:
        return pd.DataFrame()


@st.cache_data(ttl=60, show_spinner=False)
def _qualidade_por_cnpj(cnpj_norm: str) -> pd.DataFrame:
    """Busca histórico de qualidade/adimplência do CNPJ."""
    try:
        import gspread as _gs
        from google.oauth2.service_account import Credentials as _Creds
        scopes = ["https://www.googleapis.com/auth/spreadsheets",
                  "https://www.googleapis.com/auth/drive"]
        creds  = _Creds.from_service_account_info(
            dict(st.secrets["gcp_service_account"]), scopes=scopes)
        gc_q   = _gs.authorize(creds)
        sh_q   = gc_q.open_by_key(QUALIDADE_SHEET_ID)
        aba    = sh_q.worksheet(QUALIDADE_ABA)
        rows   = aba.get_all_values()
        if not rows or len(rows) < 2:
            return pd.DataFrame()
        import unicodedata as _ud
        def _nc(s):
            s = str(s).strip().lower().replace(" ","_")
            return "".join(c for c in _ud.normalize("NFD",s) if _ud.category(c)!="Mn")
        KEYS = {"cnpj","safra","cliente","adimplente"}
        hidx = 0
        for i,row in enumerate(rows):
            if {_nc(c) for c in row if str(c).strip()} & KEYS:
                hidx = i; break
        header_raw = rows[hidx]
        data_rows  = rows[hidx+1:]
        seen,header = {},[]
        for h in header_raw:
            k = _nc(h) or "col"
            if k in seen: seen[k]+=1; k=f"{k}_{seen[k]}"
            else: seen[k]=0
            header.append(k)
        n = len(header)
        data_rows = [r+[""]*(n-len(r)) if len(r)<n else r[:n] for r in data_rows]
        df = pd.DataFrame(data_rows, columns=header)
        col_cn = next((c for c in df.columns if "cnpj" in c), None)
        if not col_cn:
            return pd.DataFrame()
        df["_cn"] = df[col_cn].astype(str).str.replace(r"\D","",regex=True)
        return df[df["_cn"] == cnpj_norm].reset_index(drop=True)
    except Exception:
        return pd.DataFrame()


def _tela_dossie(gc, cnpj_norm: str, row_carteira: dict, username: str, is_admin: bool):
    """Dossiê completo do cliente."""
    razao   = str(row_carteira.get("razao_social","—"))
    fantasia= str(row_carteira.get("nome_fantasia",""))
    vend    = str(row_carteira.get("vendedor","—"))
    lider_c = str(row_carteira.get("lider","—"))
    tbp     = str(row_carteira.get("tbp","—"))
    status  = str(row_carteira.get("status","Em Atendimento"))
    telefone= str(row_carteira.get("telefone",""))
    email_c = str(row_carteira.get("email",""))
    cidade  = str(row_carteira.get("cidade",""))
    uf      = str(row_carteira.get("uf",""))
    logr    = str(row_carteira.get("logradouro",""))
    bairro  = str(row_carteira.get("bairro",""))
    ativid  = str(row_carteira.get("atividade_economica",""))
    data_r  = str(row_carteira.get("data_registro",""))
    cor     = COR_STATUS_CART.get(status,"#64748b")
    cnpj_fmt= _formatar_cnpj(cnpj_norm)

    # ── Header do dossiê ─────────────────────────────────────────
    st.markdown(f"""
    <div style="background:linear-gradient(135deg,#0d2b1a,#14532d);border-radius:14px;
                padding:20px 24px;margin-bottom:20px;border:1px solid #1a4731">
      <div style="display:flex;justify-content:space-between;align-items:start;flex-wrap:wrap;gap:10px">
        <div>
          <div style="font-size:1.2rem;font-weight:800;color:#f1f5f9">{razao}</div>
          <div style="font-size:0.8rem;color:#86efac;margin-top:3px">
            {fantasia + " · " if fantasia and fantasia!=razao else ""}{cnpj_fmt}
          </div>
          <div style="font-size:0.75rem;color:#4ade80;margin-top:4px">
            📍 {logr}{", " + bairro if bairro else ""} · {cidade}/{uf}
          </div>
          <div style="font-size:0.75rem;color:#4ade80;margin-top:2px">
            🏭 {ativid}
          </div>
        </div>
        <div style="display:flex;flex-direction:column;align-items:flex-end;gap:6px">
          <span style="background:{cor};color:#fff;padding:4px 12px;border-radius:99px;
                       font-size:0.75rem;font-weight:700">{status}</span>
          <div style="font-size:0.72rem;color:#86efac">📅 Desde {data_r[:10]}</div>
        </div>
      </div>
      <div style="margin-top:12px;display:flex;gap:20px;flex-wrap:wrap">
        <span style="font-size:0.78rem;color:#86efac">
          👤 <b style="color:#f1f5f9">{vend}</b> · 🏅 {lider_c} · 🏢 {tbp}
        </span>
        <span style="font-size:0.78rem;color:#86efac">
          📞 {telefone or "—"} · 📧 {email_c or "—"}
        </span>
      </div>
    </div>
    """, unsafe_allow_html=True)

    # ── Três seções em tabs ───────────────────────────────────────
    td_hist, td_ped, td_qual = st.tabs([
        "📞 Histórico de Contatos",
        "📦 Pedidos / Funil",
        "💳 Qualidade",
    ])

    # ── Tab: Histórico de Contatos ────────────────────────────────
    with td_hist:
        st.markdown('<p class="section-title">➕ REGISTRAR CONTATO</p>',
                    unsafe_allow_html=True)

        col_t, col_o, col_btn = st.columns([2, 4, 1])
        with col_t:
            tipo_contato = st.selectbox(
                "Tipo", TIPOS_CONTATO,
                key=f"tipo_contato_{cnpj_norm}"
            )
        with col_o:
            obs_contato = st.text_input(
                "Observação *",
                placeholder="Ex: Cliente demonstrou interesse em 5 linhas móveis...",
                key=f"obs_contato_{cnpj_norm}"
            )
        with col_btn:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("✅ Salvar", key=f"btn_contato_{cnpj_norm}",
                         type="primary", use_container_width=True):
                if not obs_contato.strip():
                    st.error("Informe a observação.")
                else:
                    ok = _salvar_contato(gc, cnpj_norm, tipo_contato,
                                         obs_contato.strip(), username)
                    if ok:
                        st.success("📋 Contato registrado!")
                        st.rerun()

        st.markdown('<p class="section-title">📋 HISTÓRICO</p>', unsafe_allow_html=True)

        df_cont = _load_contatos_cnpj(gc, cnpj_norm)

        if df_cont.empty:
            st.markdown("""
            <div style="background:#0d1f14;border:1px solid #14532d;border-radius:10px;
                        padding:20px;text-align:center;color:#4ade80">
              📭 Nenhum contato registrado ainda. Registre o primeiro acima!
            </div>
            """, unsafe_allow_html=True)
        else:
            ICONE_TIPO = {
                "📞 Ligação":"📞","🏠 Visita":"🏠","💬 WhatsApp":"💬",
                "📧 Email":"📧","📋 Reunião":"📋","📝 Outro":"📝"
            }
            for _, c_row in df_cont.iterrows():
                tipo_c = str(c_row.get("tipo",""))
                icone  = next((v for k,v in ICONE_TIPO.items() if icone_k in tipo_c
                               for icone_k in [k]), "📝")
                data_c = str(c_row.get("data",""))
                resp_c = str(c_row.get("responsavel",""))
                obs_c  = str(c_row.get("obs",""))

                st.markdown(f"""
                <div style="background:#111827;border:1px solid #1e3a5f;border-left:4px solid #3b82f6;
                            border-radius:10px;padding:12px 16px;margin-bottom:6px">
                  <div style="display:flex;justify-content:space-between;align-items:center">
                    <div>
                      <span style="font-size:0.8rem;font-weight:700;color:#93c5fd">{tipo_c}</span>
                      <span style="font-size:0.72rem;color:#475569;margin-left:10px">
                        📅 {data_c} · 👤 {resp_c}
                      </span>
                    </div>
                  </div>
                  <div style="font-size:0.78rem;color:#cbd5e1;margin-top:6px">{obs_c}</div>
                </div>
                """, unsafe_allow_html=True)

    # ── Tab: Pedidos / Funil ──────────────────────────────────────
    with td_ped:
        st.markdown('<p class="section-title">📦 PEDIDOS DO CLIENTE</p>',
                    unsafe_allow_html=True)

        df_ped = _pedidos_por_cnpj(gc, cnpj_norm)

        if df_ped.empty:
            st.markdown("""
            <div style="background:#150d2b;border:1px solid #3b1f6e;border-radius:10px;
                        padding:20px;text-align:center;color:#a78bfa">
              📭 Nenhum pedido encontrado no DadosRadar para este CNPJ.<br>
              <span style="font-size:0.8rem;color:#64748b">
                Este cliente está em prospecção — registre contatos na aba anterior.
              </span>
            </div>
            """, unsafe_allow_html=True)
        else:
            import unicodedata as _ud2
            def _nc2(s):
                s=str(s).strip().lower()
                return "".join(c for c in _ud2.normalize("NFD",s) if _ud2.category(c)!="Mn")
            col_map_r = {_nc2(c):c for c in df_ped.columns}
            col_ped   = col_map_r.get("pedido")
            col_fila  = next((c for c in df_ped.columns if "fila" in _nc2(c)),None)
            col_ac    = next((c for c in df_ped.columns if "acess" in _nc2(c)),None)
            col_preco = next((c for c in df_ped.columns if "preco" in _nc2(c) or "valor" in _nc2(c)),None)
            col_atv   = next((c for c in df_ped.columns if "ativac" in _nc2(c) or "data_at" in _nc2(c)),None)
            col_safra = next((c for c in df_ped.columns if "safra" in _nc2(c) or "mes" in _nc2(c)),None)

            # KPIs rápidos
            total_ped = len(df_ped)
            total_ac  = 0
            try:
                total_ac = int(pd.to_numeric(df_ped[col_ac], errors="coerce").sum()) if col_ac else 0
            except Exception:
                pass

            k1, k2 = st.columns(2)
            with k1:
                st.markdown(f"""<div class="kpi-mini blue">
                  <div class="kpi-label">📦 Total Pedidos</div>
                  <div class="kpi-value">{total_ped}</div></div>""",
                  unsafe_allow_html=True)
            with k2:
                st.markdown(f"""<div class="kpi-mini green">
                  <div class="kpi-label">📶 Total Acessos</div>
                  <div class="kpi-value">{total_ac}</div></div>""",
                  unsafe_allow_html=True)

            st.markdown("")

            for _, p_row in df_ped.iterrows():
                num_ped = str(p_row.get(col_ped,"—")) if col_ped else "—"
                fila    = str(p_row.get(col_fila,"—")) if col_fila else "—"
                acessos = str(p_row.get(col_ac,"")) if col_ac else ""
                preco   = str(p_row.get(col_preco,"")) if col_preco else ""
                atv     = str(p_row.get(col_atv,"")) if col_atv else ""
                safra   = str(p_row.get(col_safra,"")) if col_safra else ""

                COR_FILA = {
                    "ATIVADO":"#22c55e","APROVADO":"#16a34a",
                    "EM TRAMITACAO":"#f59e0b","PRE-VENDA":"#f59e0b",
                    "DEVOLVIDO":"#ef4444","CANCELADO":"#dc2626",
                }
                fila_u  = fila.strip().upper()
                cor_f   = next((v for k,v in COR_FILA.items() if k in fila_u),"#64748b")
                preco_f = ""
                try:
                    preco_f = f"R$ {float(str(preco).replace(',','.').replace('R$','').strip()):,.2f}"
                except Exception:
                    preco_f = preco

                st.markdown(f"""
                <div style="background:#111827;border:1px solid #1e3a5f;border-left:4px solid {cor_f};
                            border-radius:10px;padding:12px 16px;margin-bottom:6px">
                  <div style="display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:8px">
                    <div>
                      <span style="font-size:0.85rem;font-weight:700;color:#f1f5f9">Pedido #{num_ped}</span>
                      <span style="font-size:0.72rem;color:#64748b;margin-left:10px">
                        {("📅 " + atv) if atv and atv not in ("","nan") else ""}
                        {(" · 🗓 " + safra) if safra and safra not in ("","nan") else ""}
                      </span>
                    </div>
                    <span style="background:{cor_f};color:#fff;padding:2px 10px;border-radius:99px;
                                 font-size:0.7rem;font-weight:700">{fila}</span>
                  </div>
                  <div style="font-size:0.75rem;color:#94a3b8;margin-top:6px;display:flex;gap:16px">
                    {(f"📶 {acessos} acessos") if acessos and acessos not in ("","nan","0") else ""}
                    {(" · 💰 " + preco_f) if preco_f and preco_f not in ("","nan") else ""}
                  </div>
                </div>
                """, unsafe_allow_html=True)

    # ── Tab: Qualidade ────────────────────────────────────────────
    with td_qual:
        st.markdown('<p class="section-title">💳 HISTÓRICO DE ADIMPLÊNCIA</p>',
                    unsafe_allow_html=True)

        df_q = _qualidade_por_cnpj(cnpj_norm)

        if df_q.empty:
            st.info("Nenhum registro de qualidade encontrado para este cliente.")
        else:
            import unicodedata as _ud3
            def _nc3(s):
                s=str(s).strip().lower()
                return "".join(c for c in _ud3.normalize("NFD",s) if _ud3.category(c)!="Mn")
            col_map_q = {_nc3(c):c for c in df_q.columns}
            col_safra_q = col_map_q.get("safra")
            col_adim_q  = col_map_q.get("adimplente") or next(
                (c for c in df_q.columns if "adim" in _nc3(c)), None)
            col_venc_q  = col_map_q.get("vecimento") or col_map_q.get("vencimento")
            col_valor_q = next((c for c in df_q.columns if "valor" in _nc3(c)), None)
            col_obs_q   = col_map_q.get("observacoes") or col_map_q.get("obs")

            for _, q_row in df_q.iterrows():
                safra_q = str(q_row.get(col_safra_q,"")) if col_safra_q else ""
                adim_q  = str(q_row.get(col_adim_q,"")).strip().upper() if col_adim_q else ""
                venc_q  = str(q_row.get(col_venc_q,"")).strip() if col_venc_q else ""
                valor_q = str(q_row.get(col_valor_q,"")).strip() if col_valor_q else ""
                obs_q   = str(q_row.get(col_obs_q,"")).strip()[:150] if col_obs_q else ""

                adim_key_q = adim_q
                if "N" in adim_key_q and adim_key_q not in ("SIM","FATURA GERADA"):
                    adim_key_q = "NAO"

                cor_q  = COR_ADIMPLENTE.get(adim_key_q, "#64748b")
                icon_q = ICON_ADIMPLENTE.get(adim_key_q, "▪️")

                preco_q = ""
                try:
                    preco_q = f"R$ {float(valor_q.split('/')[0].replace(',','.').replace('R$','').strip()):,.2f}"
                except Exception:
                    preco_q = valor_q

                st.markdown(f"""
                <div style="background:#111827;border:1px solid #1e3a5f;border-left:4px solid {cor_q};
                            border-radius:10px;padding:12px 16px;margin-bottom:6px">
                  <div style="display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:8px">
                    <div>
                      <span style="font-size:0.85rem;font-weight:700;color:#f1f5f9">Safra {safra_q}</span>
                      {(f'<span style="font-size:0.72rem;color:#64748b;margin-left:10px">Vence: {venc_q}</span>') if venc_q else ""}
                    </div>
                    <div style="display:flex;align-items:center;gap:8px">
                      {(f'<span style="color:#f59e0b;font-weight:700;font-size:0.82rem">{preco_q}</span>') if preco_q and preco_q not in ("","nan") else ""}
                      <span style="background:{cor_q};color:#fff;padding:2px 10px;border-radius:99px;
                                   font-size:0.7rem;font-weight:700">{icon_q} {adim_q or "—"}</span>
                    </div>
                  </div>
                  {(f'<div style="font-size:0.75rem;color:#6b7280;margin-top:6px">{obs_q}</div>') if obs_q else ""}
                </div>
                """, unsafe_allow_html=True)


def _tela_carteira_novo(gc, df_carteira, vendedores, mapa_lider, mapa_tbp, username, info):
    st.markdown('<p class="section-title">🔍 BUSCAR CLIENTE PELO CNPJ</p>', unsafe_allow_html=True)
    col_ci, col_cb = st.columns([3,1])
    with col_ci:
        cnpj_in = st.text_input("CNPJ do cliente", placeholder="00.000.000/0000-00", key="cart_cnpj_in")
    with col_cb:
        st.markdown("<br>", unsafe_allow_html=True)
        buscar = st.button("🔍 Consultar Receita", use_container_width=True, key="cart_btn_buscar")

    if buscar and cnpj_in:
        cnpj_limpo = re.sub(r"\D","",cnpj_in)
        df_atual = _atualizar_expirados_gc(gc, df_carteira.copy())
        disponivel, row_existe = _cnpj_disponivel(df_atual, cnpj_limpo)

        if not disponivel and row_existe:
            vb  = row_existe.get("vendedor","—")
            lb  = row_existe.get("lider","—")
            tb  = row_existe.get("tbp","—")
            db  = row_existe.get("data_registro","")
            dr  = _dias_restantes_cart(db)
            st.markdown(f"""
            <div style="background:#1c0a0a;border:1px solid #7f1d1d;border-radius:10px;
                        padding:16px 20px;margin:12px 0;color:#fca5a5">
              <div style="font-size:1rem;font-weight:800;color:#f87171;margin-bottom:8px">🔒 Cliente ja em atendimento</div>
              <div><b>Vendedor:</b> {vb} &nbsp;·&nbsp; <b>Lider:</b> {lb} &nbsp;·&nbsp; <b>Escritorio:</b> {tb}</div>
              <div style="margin-top:4px"><b>Desde:</b> {db[:10]} &nbsp;·&nbsp; <b>Expira em:</b> {dr} dias</div>
              <div style="margin-top:10px;font-size:0.82rem">⚠️ Para liberar, entre em contato com o <b>administrador</b>.</div>
            </div>
            """, unsafe_allow_html=True)
            return

        with st.spinner("Consultando Receita Federal..."):
            dados_rec = _buscar_cnpj_receita(cnpj_limpo)
        if "erro" in dados_rec:
            st.error(f"❌ {dados_rec['erro']}")
            return
        if row_existe:
            st.info(f"ℹ️ CNPJ ja esteve na carteira com status **{row_existe.get('status','')}**. Pode registrar novamente.")
        st.success(f"✅ **{dados_rec['razao_social']}** encontrado!")
        st.session_state["cart_dados_receita"] = dados_rec

    c = st.session_state.get("cart_dados_receita", {})
    if not c:
        st.markdown("""
        <div style="background:#0d1f14;border:1px solid #14532d;border-radius:10px;
                    padding:24px;text-align:center;margin-top:20px">
          <div style="font-size:2.5rem">📁</div>
          <div style="color:#4ade80;font-weight:700;margin-top:8px">Digite o CNPJ para iniciar o atendimento</div>
          <div style="color:#64748b;font-size:0.82rem;margin-top:4px">Dados preenchidos automaticamente pela Receita Federal</div>
        </div>
        """, unsafe_allow_html=True)
        return

    # ── Responsável FORA do form para reagir ao selectbox ────────
    st.markdown('<p class="section-title">👤 RESPONSÁVEL PELO ATENDIMENTO</p>', unsafe_allow_html=True)
    cv, cl, ct = st.columns(3)
    with cv:
        if vendedores:
            vend_sel = st.selectbox("Vendedor *", vendedores, key="cart_form_vend")
        else:
            st.warning("⚠️ Nenhum vendedor com líder encontrado na aba Colaboradores.")
            vend_sel = st.text_input("Vendedor *", key="cart_form_vend_txt")

    # Preenche líder e TBP automaticamente — reage a cada mudança
    lider_auto = mapa_lider.get(str(vend_sel), "") if vend_sel else ""
    tbp_auto   = mapa_tbp.get(str(vend_sel), "")   if vend_sel else ""

    with cl:
        st.markdown(f"""
        <div style="background:#0d1f14;border:1px solid #14532d;border-radius:8px;
                    padding:10px 14px;min-height:46px">
          <div style="font-size:0.65rem;color:#4ade80;font-weight:700;text-transform:uppercase;
                      letter-spacing:0.5px;margin-bottom:4px">Líder</div>
          <div style="font-size:0.92rem;font-weight:700;color:#93c5fd">
            {lider_auto if lider_auto else "<span style='color:#475569'>—</span>"}
          </div>
        </div>
        """, unsafe_allow_html=True)
    with ct:
        st.markdown(f"""
        <div style="background:#150d2b;border:1px solid #3b1f6e;border-radius:8px;
                    padding:10px 14px;min-height:46px">
          <div style="font-size:0.65rem;color:#a78bfa;font-weight:700;text-transform:uppercase;
                      letter-spacing:0.5px;margin-bottom:4px">Escritório (TBP)</div>
          <div style="font-size:0.92rem;font-weight:700;color:#a78bfa">
            {tbp_auto if tbp_auto else "<span style='color:#475569'>—</span>"}
          </div>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("")

    def _sv(d, k):
        v = d.get(k, "")
        return str(v) if v is not None else ""

    with st.form("form_carteira_novo"):
        st.markdown('<p class="section-title">🏢 DADOS DO CLIENTE</p>', unsafe_allow_html=True)
        c1, c2, c3 = st.columns(3)
        with c1:
            razao    = st.text_input("Razao Social *",  value=_sv(c,"razao_social"))
            telefone = st.text_input("Telefone *",       value=_sv(c,"telefone"))
            cidade   = st.text_input("Cidade",          value=_sv(c,"cidade"))
        with c2:
            fantasia  = st.text_input("Nome Fantasia",  value=_sv(c,"nome_fantasia"))
            email_cli = st.text_input("Email *",         value=_sv(c,"email"))
            uf        = st.text_input("UF",             value=_sv(c,"uf"))
        with c3:
            atividade = st.text_input("Atividade",      value=_sv(c,"atividade_economica"))
            fundacao  = st.text_input("Data Fundacao",  value=_sv(c,"data_fundacao"))
            capital   = st.text_input("Capital Social", value=_sv(c,"capital_social"))

        obs_val = st.text_area("Observacoes", height=60, key="cart_form_obs")

        submitted = st.form_submit_button("📁 Registrar na Carteira", type="primary", use_container_width=True)
        if submitted:
            vend_final  = str(st.session_state.get("cart_form_vend", vend_sel)).strip()
            lider_final = mapa_lider.get(vend_final, "")
            tbp_final   = mapa_tbp.get(vend_final, "")
            if not razao.strip():
                st.error("⚠️ Razão Social obrigatória.")
            elif not telefone.strip():
                st.error("⚠️ Telefone de contato é obrigatório.")
            elif not email_cli.strip():
                st.error("⚠️ Email de contato é obrigatório.")
            elif not vend_final:
                st.error("⚠️ Selecione o vendedor.")
            else:
                dados_salvar = {
                    "cnpj":                re.sub(r"\D","",c.get("cnpj","")),
                    "razao_social":        razao.strip(),
                    "nome_fantasia":       fantasia.strip(),
                    "atividade_economica": atividade.strip(),
                    "data_fundacao":       fundacao.strip(),
                    "capital_social":      capital.strip(),
                    "telefone":            telefone.strip(),
                    "email":               email_cli.strip(),
                    "cep":                 c.get("cep",""),
                    "logradouro":          c.get("logradouro",""),
                    "numero":              c.get("numero",""),
                    "bairro":              c.get("bairro",""),
                    "cidade":              cidade.strip(),
                    "uf":                  uf.strip(),
                    "vendedor":            vend_final,
                    "lider":               lider_final,
                    "tbp":                 tbp_final,
                    "obs":                 obs_val.strip(),
                    "pedido_tim":          "",
                }
                with st.spinner("Registrando..."):
                    ok = _registrar_cnpj_carteira(gc, dados_salvar, username)
                if ok:
                    st.success(f"✅ **{razao}** registrado! Vendedor: **{vend_final}** · Líder: **{lider_final}** · Escritório: **{tbp_final}**")
                    st.session_state.pop("cart_dados_receita", None)
                    st.rerun()


def _tela_carteira_lista(gc, df, info, username, is_admin):
    tipo    = info.get("tipo","lider")
    lider_u = info.get("lider","")

    st.markdown('<p class="section-title">🔍 FILTROS</p>', unsafe_allow_html=True)
    cf1, cf2, cf3, cf4 = st.columns(4)
    with cf1:
        busca = st.text_input("Empresa / CNPJ", key="cart_lst_busca")
    with cf2:
        sf = st.selectbox("Status", ["Todos"]+STATUS_CARTEIRA, key="cart_lst_sf")
    with cf3:
        if is_admin and "tbp" in df.columns:
            tbp_f = st.selectbox("Escritorio (TBP)", ["Todos"]+sorted(df["tbp"].dropna().unique().tolist()), key="cart_lst_tbp")
        else:
            tbp_f = "Todos"
    with cf4:
        if is_admin and "lider" in df.columns:
            lider_f = st.selectbox("Lider", ["Todos"]+sorted(df["lider"].dropna().unique().tolist()), key="cart_lst_lf")
        else:
            lider_f = "Todos"

    df_f = df.copy()
    if not is_admin and tipo == "lider" and lider_u:
        df_f = df_f[df_f["lider"].astype(str).str.strip().str.upper() == lider_u.strip().upper()]
    if busca:
        mask = (df_f["razao_social"].astype(str).str.contains(busca,case=False,na=False) |
                df_f["cnpj"].astype(str).str.contains(re.sub(r"\D","",busca),na=False))
        df_f = df_f[mask]
    if sf != "Todos":
        df_f = df_f[df_f["status"] == sf]
    if tbp_f != "Todos" and "tbp" in df_f.columns:
        df_f = df_f[df_f["tbp"].astype(str).str.strip() == tbp_f]
    if lider_f != "Todos" and "lider" in df_f.columns:
        df_f = df_f[df_f["lider"].astype(str).str.strip() == lider_f]
    if "data_registro" in df_f.columns:
        df_f = df_f.sort_values("data_registro", ascending=False)

    n_total = len(df_f)
    n_atend = len(df_f[df_f["status"]=="Em Atendimento"])
    n_ativ  = len(df_f[df_f["status"]=="Ativado"])
    n_exp   = len(df_f[df_f["status"]=="Expirado"])

    k1,k2,k3,k4 = st.columns(4)
    for col_k, val, label, sub, cor in [
        (k1, n_total, "📁 Total",          "na carteira",          "blue"),
        (k2, n_atend, "⏳ Em Atendimento", "em andamento",          "amber"),
        (k3, n_ativ,  "✅ Ativados",        "fechados",              "green"),
        (k4, n_exp,   "💤 Expirados",       f"{DIAS_EXPIRACAO}+ d", "purple"),
    ]:
        with col_k:
            st.markdown(f"""<div class="kpi-mini {cor}">
              <div class="kpi-label">{label}</div>
              <div class="kpi-value">{val}</div>
              <div class="kpi-sub">{sub}</div></div>""", unsafe_allow_html=True)

    st.markdown(f"<br><span style='color:#64748b;font-size:0.8rem'>{n_total} cliente(s)</span>", unsafe_allow_html=True)

    if df_f.empty:
        st.markdown("""
        <div style="background:#0d1f14;border:1px solid #14532d;border-radius:10px;
                    padding:24px;text-align:center;margin-top:16px">
          <div style="font-size:2rem">📭</div>
          <div style="color:#4ade80;font-weight:700;margin-top:8px">Nenhum cliente encontrado</div>
          <div style="color:#64748b;font-size:0.82rem">Use <b>➕ Novo Registro</b> para adicionar</div>
        </div>
        """, unsafe_allow_html=True)
        return

    for _, row in df_f.iterrows():
        _card_cliente_carteira(row.to_dict(), is_admin, gc, username)

    if is_admin:
        st.markdown('<p class="section-title">🔓 LIBERAR CNPJ</p>', unsafe_allow_html=True)
        cl1, cl2, cl3 = st.columns([2,3,1])
        with cl1:
            cnpj_lib = st.text_input("CNPJ", placeholder="00.000.000/0000-00", key="cart_lib_cnpj")
        with cl2:
            mot_lib  = st.text_input("Motivo", key="cart_lib_mot")
        with cl3:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("🔓 Liberar", type="primary", use_container_width=True, key="cart_lib_btn"):
                if cnpj_lib.strip():
                    ok = _atualizar_campo_carteira(gc, cnpj_lib.strip(), {
                        "status": "Liberado pelo Admin",
                        "obs": f"[Admin] {mot_lib.strip()}" if mot_lib.strip() else "[Admin] Liberado manualmente",
                    })
                    if ok:
                        st.success(f"✅ {_formatar_cnpj(cnpj_lib)} liberado!")
                        st.rerun()
                    else:
                        st.error("CNPJ nao encontrado.")
                else:
                    st.error("Informe o CNPJ.")
        st.markdown("")
        st.download_button("⬇️ Exportar CSV", df_f.to_csv(index=False).encode("utf-8"),
                           "carteira.csv", "text/csv", key="cart_exp")




# ─────────────────────────────────────────────────────────────────
#  QUALIDADE — carregamento e render
# ─────────────────────────────────────────────────────────────────

QUALIDADE_SHEET_ID = "1QO4EpMnVFbHMMGASzjhmemlC2avuyhZ8bHIcdSnAnlU"
QUALIDADE_ABA      = "BASE_SAFRAS_QUALIDADE"

COR_ADIMPLENTE = {
    "SIM":           "#22c55e",
    "NAO":           "#ef4444",
    "FATURA GERADA": "#f59e0b",
}
ICON_ADIMPLENTE = {
    "SIM":           "\u2705",
    "NAO":           "\u274c",
    "FATURA GERADA": "\u23f3",
}


@st.cache_data(ttl=60, show_spinner=False)
def _load_qualidade():
    import unicodedata, gspread
    from google.oauth2.service_account import Credentials
    def _nq(s):
        s = str(s).strip().lower()
        return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    try:
        scopes = ["https://www.googleapis.com/auth/spreadsheets",
                  "https://www.googleapis.com/auth/drive"]
        creds  = Credentials.from_service_account_info(
            dict(st.secrets["gcp_service_account"]), scopes=scopes)
        gc_q   = gspread.authorize(creds)
        aba    = gc_q.open_by_key(QUALIDADE_SHEET_ID).worksheet(QUALIDADE_ABA)
        rows   = aba.get_all_values()
        if not rows or len(rows) < 2:
            return pd.DataFrame()

        KEYWORDS = {"cnpj", "safra", "cliente", "adimplente"}
        header_idx = 0
        for i, row in enumerate(rows):
            if {_nq(c) for c in row if str(c).strip()} & KEYWORDS:
                header_idx = i
                break

        header_raw = rows[header_idx]
        data_rows  = rows[header_idx + 1:]
        n = len(header_raw)

        seen, header = {}, []
        for h in header_raw:
            k = _nq(h) or "col"
            if k in seen:
                seen[k] += 1; k = f"{k}_{seen[k]}"
            else:
                seen[k] = 0
            header.append(k)

        rows_pad = [r + [""] * (n - len(r)) if len(r) < n else r[:n] for r in data_rows]
        df = pd.DataFrame(rows_pad, columns=header)
        df = df[df.apply(lambda r: any(str(v).strip() for v in r), axis=1)].copy()

        col_cnpj = next((c for c in df.columns if "cnpj" in c), None)
        if col_cnpj:
            df["cnpj_norm"] = (df[col_cnpj].astype(str)
                               .str.replace(r"\D", "", regex=True).str.strip())
        else:
            df["cnpj_norm"] = ""
        return df
    except Exception as e:
        st.error(f"Erro ao carregar Qualidade: {e}")
        return pd.DataFrame()


def _cruzar_qualidade_bko(df_qual: pd.DataFrame, df_bko: pd.DataFrame) -> pd.DataFrame:
    """
    Cruza qualidade com BKO via DadosRadar como intermediário:
    Qualidade(CNPJ) → DadosRadar(CNPJ→pedido) → BKO(pedido→vendedor+líder)
    """
    import unicodedata, gspread
    def _nq(s):
        s = str(s).strip().lower()
        return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    def _np(v):
        s = str(v).strip()
        if s.endswith(".0"): s = s[:-2]
        return s.lstrip("0") or s

    df_qual["vendedor_real"] = ""
    df_qual["lider_bko"]    = ""

    if df_qual.empty or "cnpj_norm" not in df_qual.columns:
        return df_qual

    try:
        # ── 1. DadosRadar: CNPJ → pedido ─────────────────────────
        from google.oauth2.service_account import Credentials
        scopes = ["https://www.googleapis.com/auth/spreadsheets",
                  "https://www.googleapis.com/auth/drive"]
        creds  = Credentials.from_service_account_info(
            dict(st.secrets["gcp_service_account"]), scopes=scopes)
        gc_r   = gspread.authorize(creds)
        sh     = gc_r.open_by_key(SPREADSHEET_ID)
        radar  = sh.worksheet("DadosRadar")
        r_rows = radar.get_all_values()

        mapa_cnpj_pedido = {}
        if r_rows and len(r_rows) > 1:
            r_header = [str(h).strip().lower() for h in r_rows[0]]
            ci_cn = next((i for i,h in enumerate(r_header) if h == "cnpj"), None)
            ci_pd = next((i for i,h in enumerate(r_header) if h == "pedido"), None)
            if ci_cn is not None and ci_pd is not None:
                for row in r_rows[1:]:
                    if not row or len(row) <= max(ci_cn, ci_pd):
                        continue
                    cn = re.sub(r"\D","",str(row[ci_cn]).strip())
                    pd_ = _np(str(row[ci_pd]))
                    if cn and pd_ and len(cn) == 14:
                        mapa_cnpj_pedido[cn] = pd_

        # ── 2. BKO-VENDEDOR-REAL: pedido → vendedor + líder ──────
        col_map = {_nq(c): c for c in df_bko.columns}
        col_ped = col_map.get("pedido")
        col_v   = col_map.get("vendedor_real") or next(
            (c for c in df_bko.columns if "vendedor" in _nq(c)), None)
        col_l   = col_map.get("lider")

        mapa_ped_vend = {}
        if col_ped and col_v and not df_bko.empty:
            for _, row in df_bko.iterrows():
                ped = _np(str(row.get(col_ped, "")))
                vd  = str(row.get(col_v, "")).strip()
                ld  = str(row.get(col_l, "")).strip() if col_l else ""
                if ped and vd and vd.lower() not in ("","nan","none","sem vendedor"):
                    mapa_ped_vend[ped] = {"vendedor_real": vd, "lider_bko": ld}

        # ── 3. Aplica o cruzamento ────────────────────────────────
        def _get_vendedor(cnpj):
            ped  = mapa_cnpj_pedido.get(cnpj, "")
            info = mapa_ped_vend.get(ped, {})
            return info.get("vendedor_real", "")

        def _get_lider(cnpj):
            ped  = mapa_cnpj_pedido.get(cnpj, "")
            info = mapa_ped_vend.get(ped, {})
            return info.get("lider_bko", "")

        df_qual["vendedor_real"] = df_qual["cnpj_norm"].apply(_get_vendedor)
        df_qual["lider_bko"]    = df_qual["cnpj_norm"].apply(_get_lider)

    except Exception as e:
        st.warning(f"Qualidade: erro no cruzamento — {e}")

    return df_qual


def _tela_qualidade(df_bko: pd.DataFrame, info: dict, tipo: str, lider_u: str):
    import unicodedata
    def _nq(s):
        s = str(s).strip().lower()
        return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")

    st.markdown('<p class="section-title">\U0001f4ca QUALIDADE \u2014 STATUS DE PAGAMENTO</p>',
                unsafe_allow_html=True)

    with st.spinner("Carregando dados de qualidade..."):
        df_qual = _cruzar_qualidade_bko(_load_qualidade(), df_bko)

    if df_qual.empty:
        st.warning("\u26a0\ufe0f Nenhum dado encontrado na planilha de qualidade.")
        return

    col_map     = {_nq(c): c for c in df_qual.columns}
    col_safra   = col_map.get("safra")
    col_cnpj    = col_map.get("cnpj")
    col_cliente = col_map.get("cliente")
    col_adim    = col_map.get("adimplente") or next(
        (c for c in df_qual.columns if "adim" in _nq(c)), None)
    col_vencto  = col_map.get("vecimento") or col_map.get("vencimento")
    col_valor   = next((c for c in df_qual.columns if "valor" in _nq(c)), None)
    col_obs     = col_map.get("observacoes") or col_map.get("obs")
    col_contato = next((c for c in df_qual.columns
                        if "contato" in _nq(c) and "cliente" in _nq(c)), None)
    col_parceiro = col_map.get("parceiro")

    is_admin = (tipo == "admin")

    # ── Aplica filtro de perfil primeiro (antes de montar listas) ─
    df_f = df_qual.copy()
    if not is_admin:
        if tipo == "lider" and lider_u:
            df_f = df_f[df_f["lider_bko"].astype(str).str.strip().str.upper()
                        == lider_u.strip().upper()]
        else:
            nome_u = info.get("name", "").strip().upper()
            df_f   = df_f[df_f["vendedor_real"].astype(str).str.strip().str.upper() == nome_u]

    # ── Filtros de UI ─────────────────────────────────────────────
    if is_admin:
        cf1, cf2, cf3, cf4, cf5, cf6 = st.columns(6)
    else:
        cf1, cf2, cf3, cf4 = st.columns(4)
        cf5 = cf6 = None

    with cf1:
        safras = ["Todas"]
        if col_safra:
            safras += sorted(df_qual[col_safra].dropna().unique().tolist(), reverse=True)
        safra_f = st.selectbox("📅 Safra", safras, key="qual_safra")

    with cf2:
        adim_f = st.selectbox("💳 Adimplente?",
                              ["Todos", "NAO", "SIM", "FATURA GERADA"], key="qual_adim")

    with cf3:
        if is_admin:
            lideres_u = ["Todos"] + sorted(
                v for v in df_qual["lider_bko"].dropna().unique()
                if v and v not in ("", "nan"))
            lider_f = st.selectbox("🏅 Líder", lideres_u, key="qual_lider_f")
        else:
            lider_f = "Todos"
            busca = st.text_input("🔍 Empresa / CNPJ", key="qual_busca")

    with cf4:
        if is_admin:
            vends_u = ["Todos"] + sorted(
                v for v in df_qual["vendedor_real"].dropna().unique()
                if v and v not in ("", "nan"))
            vend_f = st.selectbox("👤 Vendedor", vends_u, key="qual_vend")
        else:
            vend_f = "Todos"

    if is_admin and cf5:
        with cf5:
            parcs_u = ["Todos"]
            if col_parceiro and col_parceiro in df_qual.columns:
                parcs_u += sorted(
                    v for v in df_qual[col_parceiro].dropna().unique()
                    if v and v not in ("", "nan"))
            parc_f = st.selectbox("🏢 Parceiro (TBP)", parcs_u, key="qual_parc_f")
        with cf6:
            busca = st.text_input("🔍 Empresa / CNPJ", key="qual_busca")
    else:
        parc_f = "Todos"
        if not cf5:  # não-admin já definiu busca no cf3
            pass
        else:
            busca = ""

    # ── Aplica filtros de UI ──────────────────────────────────────
    if safra_f != "Todas" and col_safra:
        df_f = df_f[df_f[col_safra].astype(str).str.strip() == safra_f]
    if adim_f != "Todos" and col_adim:
        df_f = df_f[df_f[col_adim].astype(str).str.strip().str.upper()
                    .str.replace("NÃO","NAO").str.replace("N\u00c3O","NAO") == adim_f]
    if lider_f != "Todos":
        df_f = df_f[df_f["lider_bko"].astype(str).str.strip() == lider_f]
    if vend_f != "Todos":
        df_f = df_f[df_f["vendedor_real"].astype(str).str.strip() == vend_f]
    if parc_f != "Todos" and col_parceiro and col_parceiro in df_f.columns:
        df_f = df_f[df_f[col_parceiro].astype(str).str.strip() == parc_f]
    if busca:
        mask = pd.Series(False, index=df_f.index)
        if col_cliente:
            mask |= df_f[col_cliente].astype(str).str.contains(busca, case=False, na=False)
        if col_cnpj:
            mask |= df_f[col_cnpj].astype(str).str.contains(
                re.sub(r"\D","",busca), na=False)
        df_f = df_f[mask]

    if col_cliente:
        df_f = df_f[df_f[col_cliente].astype(str).str.strip() != ""]

    n_total   = len(df_f)
    n_inad    = len(df_f[df_f[col_adim].astype(str).str.strip().str.upper()
                         .str.contains("N", na=False)]) if col_adim else 0
    n_ok      = len(df_f[df_f[col_adim].astype(str).str.strip().str.upper() == "SIM"]) if col_adim else 0
    n_sem_v   = len(df_f[df_f["vendedor_real"].astype(str).str.strip() == ""])

    k1,k2,k3,k4 = st.columns(4)
    for col_k, val, label, sub, cor in [
        (k1, n_total, "\U0001f4cb Total",          "registros",           "blue"),
        (k2, n_inad,  "\u274c Inadimplentes",       "fatura em aberto",    "red"),
        (k3, n_ok,    "\u2705 Adimplentes",          "em dia",              "green"),
        (k4, n_sem_v, "\u26a0\ufe0f Sem Vendedor",  "nao cruzados no BKO", "amber"),
    ]:
        with col_k:
            st.markdown(f'''<div class="kpi-mini {cor}">
              <div class="kpi-label">{label}</div>
              <div class="kpi-value">{val}</div>
              <div class="kpi-sub">{sub}</div></div>''', unsafe_allow_html=True)

    st.markdown(f"<br><span style='color:#64748b;font-size:0.8rem'>{n_total} registro(s)</span>",
                unsafe_allow_html=True)

    if df_f.empty:
        st.info("Nenhum registro encontrado para os filtros selecionados.")
        return

    # Ordena: inadimplentes primeiro
    if col_adim:
        df_f = df_f.sort_values(col_adim, ascending=True, key=lambda s: s.astype(str))

    def _esc(v, maxlen=None):
        s = str(v).strip()
        if maxlen:
            s = s[:maxlen]
        return (s.replace("&","&amp;").replace("<","&lt;").replace(">","&gt;")
                 .replace('"',"&quot;").replace("'","&#39;"))

    for _, row in df_f.iterrows():
        cliente  = _esc(row.get(col_cliente,"") if col_cliente else "")  or "—"
        cnpj_raw = str(row.get(col_cnpj,"")) if col_cnpj else ""
        cnpj_fmt = _esc(_formatar_cnpj(cnpj_raw)) if cnpj_raw else "—"
        adim_raw = str(row.get(col_adim,"")).strip().upper() if col_adim else ""
        adim_key = adim_raw
        if "N" in adim_key and adim_key not in ("SIM","FATURA GERADA"):
            adim_key = "NAO"
        vencto   = _esc(row.get(col_vencto,"") if col_vencto else "")
        valor    = _esc(row.get(col_valor,"") if col_valor else "")
        obs      = _esc(row.get(col_obs,"") if col_obs else "", maxlen=200)
        safra    = _esc(row.get(col_safra,"") if col_safra else "")
        parceiro = _esc(row.get(col_parceiro,"") if col_parceiro else "")
        vendedor = _esc(row.get("vendedor_real",""))
        lider_c  = _esc(row.get("lider_bko",""))
        adim_lbl = _esc(adim_raw) if adim_raw else "—"

        cor_c  = COR_ADIMPLENTE.get(adim_key, "#64748b")
        icon_c = ICON_ADIMPLENTE.get(adim_key, "▪️")

        valor_html  = (f'<span style="color:#f59e0b;font-weight:700">R$ {valor}</span>'
                       f' &nbsp;·&nbsp;') if valor and valor not in ("","nan") else ""
        vencto_html = (f'<span style="color:#94a3b8;font-size:0.72rem">Vence: {vencto}</span>'
                      ) if vencto and vencto not in ("","nan") else ""
        obs_html    = (f'<div style="font-size:0.72rem;color:#6b7280;margin-top:6px;'
                       f'border-left:2px solid #334155;padding-left:8px">{obs}</div>'
                      ) if obs and obs not in ("","nan") else ""
        vend_html   = (f'👤 <b style="color:#e2e8f0">{vendedor}</b> &nbsp;·&nbsp; 🏅 {lider_c}'
                       if vendedor else
                       '<span style="color:#ef4444;font-size:0.72rem">⚠️ Sem vendedor no BKO</span>')

        html_card = (
            f'<div style="background:#111827;border:1px solid #1e3a5f;'
            f'border-left:5px solid {cor_c};border-radius:12px;'
            f'padding:14px 18px;margin-bottom:8px">'
            f'<div style="display:flex;justify-content:space-between;'
            f'align-items:start;flex-wrap:wrap;gap:8px">'
            f'<div style="flex:1;min-width:200px">'
            f'<div style="font-size:0.95rem;font-weight:700;color:#f1f5f9">{cliente}</div>'
            f'<div style="font-size:0.73rem;color:#64748b;margin-top:2px">'
            f'{cnpj_fmt} · 📅 {safra} · 🏢 {parceiro}</div>'
            f'<div style="font-size:0.73rem;color:#64748b;margin-top:5px">{vend_html}</div>'
            f'<div style="margin-top:6px;display:flex;align-items:center;'
            f'flex-wrap:wrap;gap:4px">{valor_html}{vencto_html}</div>'
            f'{obs_html}'
            f'</div>'
            f'<div style="display:flex;flex-direction:column;align-items:flex-end;gap:6px">'
            f'<span style="display:inline-block;padding:3px 10px;border-radius:99px;'
            f'font-size:0.7rem;font-weight:700;color:#fff;background:{cor_c}">'
            f'{icon_c} {adim_lbl}</span>'
            f'</div></div></div>'
        )
        st.markdown(html_card, unsafe_allow_html=True)

    st.markdown("")
    cols_exp = [c for c in [col_safra, col_parceiro, col_cnpj, col_cliente,
                              col_adim, col_vencto, col_valor,
                              "vendedor_real", "lider_bko", col_obs]
                if c and c in df_f.columns]
    st.download_button("\u2b07\ufe0f Exportar CSV",
                       df_f[cols_exp].to_csv(index=False).encode("utf-8"),
                       "qualidade.csv", "text/csv", key="qual_export")

# ─────────────────────────────────────────────────────────────────
#  CARREGAMENTO DE DADOS
# ─────────────────────────────────────────────────────────────────

@st.cache_data(ttl=600, show_spinner=False)
def _filtrar_qualidade_por_perfil_local(df: pd.DataFrame, tipo: str,
                                         lider_u: str, parceiro_u: str) -> pd.DataFrame:
    if df.empty:
        return df
    if tipo == "admin":
        return df
    if tipo == "lider" and lider_u:
        col = next((c for c in df.columns if "lider" in c.lower()), None)
        if col:
            return df[df[col].astype(str).str.strip().str.upper() == lider_u.strip().upper()]
    if tipo == "parceiro" and parceiro_u:
        col = next((c for c in df.columns if "vendedor" in c.lower()), None)
        if col:
            return df[df[col].astype(str).str.strip().str.upper() == parceiro_u.strip().upper()]
    return df


def _load_all():
    raw   = load_data()
    bko   = load_bko()
    colab = load_colaboradores()
    return raw, bko, colab


@st.cache_data(ttl=300, show_spinner=False)
def _carregar_pendentes_bko(lider_filtro="", parceiro_filtro="", tipo="admin"):
    gc         = get_gspread_client()
    planilha   = gc.open_by_key(SPREADSHEET_ID)
    ws         = planilha.worksheet("BKO-VENDEDOR-REAL")
    all_values = ws.get_all_values()
    if not all_values or len(all_values) < 3:
        return pd.DataFrame(), ws

    headers = all_values[1]
    rows    = all_values[2:]
    df = pd.DataFrame(rows, columns=headers)
    df.columns = [c.strip().lower().replace(" ", "_") for c in df.columns]
    df["_row_idx"] = [i + 3 for i in range(len(df))]

    col_pedido   = next((c for c in df.columns if "pedido" in c), None)
    col_razao    = next((c for c in df.columns if "raz" in c and "social" in c), None)
    col_vendedor = next((c for c in df.columns if "vendedor" in c and "real" in c), None)
    col_safra    = next((c for c in df.columns if "safra" in c), None)
    col_tipo     = next((c for c in df.columns if "tipo" in c and "contrat" in c), None)
    col_fila     = next((c for c in df.columns if "fila" in c), None)

    if not col_pedido or not col_vendedor:
        return pd.DataFrame(), ws

    df_pend = df[df[col_vendedor].apply(lambda x: _s(x) == "")].copy()
    df_pend = df_pend[df_pend[col_pedido].apply(lambda x: _norm_pedido(x) != "")].copy()

    if col_tipo:
        df_pend = df_pend[~df_pend[col_tipo].apply(lambda x: "RENEGOCI" in _s(x).upper())]
    if col_fila:
        df_pend = df_pend[~df_pend[col_fila].apply(lambda x: "CANCELAD" in _s(x).upper())]

    from datetime import timedelta
    limite = datetime.today() - timedelta(days=60)
    if col_safra:
        def _parse(v):
            try:
                return datetime.strptime(_s(v).strip(), "%m/%Y")
            except Exception:
                return None
        df_pend["_dt"] = df_pend[col_safra].apply(_parse)
        df_pend = df_pend[df_pend["_dt"].apply(lambda d: d is not None and d >= limite.replace(day=1))]

    rename = {}
    if col_pedido:   rename[col_pedido]   = "pedido"
    if col_razao:    rename[col_razao]    = "razao_social"
    if col_safra:    rename[col_safra]    = "safra"
    if col_vendedor: rename[col_vendedor] = "vendedor_real"

    df_pend = df_pend.rename(columns=rename)
    if "pedido" not in df_pend.columns:
        return pd.DataFrame(), ws
    df_pend["pedido"] = df_pend["pedido"].apply(_norm_pedido)
    df_pend = df_pend[df_pend["pedido"] != ""].drop_duplicates("pedido").reset_index(drop=True)

    try:
        col_vend_idx = headers.index(
            next(h for h in headers if "vendedor" in h.lower() and "real" in h.lower())
        ) + 1
    except Exception:
        col_vend_idx = 4

    df_pend["_col_vendedor_idx"] = col_vend_idx
    colunas = [c for c in ["pedido","razao_social","safra","_row_idx","_col_vendedor_idx"] if c in df_pend.columns]
    return df_pend[colunas], ws


# ─────────────────────────────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────────────────────────────

def main():
    # ── Auth ──────────────────────────────────────────────────────
    auth_config   = _carregar_auth()
    authenticator = stauth.Authenticate(
        auth_config["credentials"],
        auth_config["cookie"]["name"],
        auth_config["cookie"]["key"],
        auth_config["cookie"]["expiry_days"],
    )

    login_result = authenticator.login(location="main")
    if login_result:
        name, authentication_status, username = login_result
    else:
        name                  = st.session_state.get("name")
        authentication_status = st.session_state.get("authentication_status")
        username              = st.session_state.get("username")

    if authentication_status is False:
        st.error("Usuário ou senha incorretos.")
        st.stop()
    if authentication_status is None:
        st.info("Informe seu usuário e senha para acessar.")
        st.stop()

    # ── Usuário logado ────────────────────────────────────────────
    info       = _info_usuario(username)
    tipo       = info["tipo"]
    lider_u    = info["lider"]
    parceiro_u = info["parceiro"]

    # ── Header ────────────────────────────────────────────────────
    mes_str = MESES_PT.get(datetime.now().strftime("%m"), "") + "/" + datetime.now().strftime("%Y")
    badge   = "🔑 ADMIN" if tipo == "admin" else ("👤 LÍDER" if tipo == "lider" else "🏢 PARCEIRO")
    st.markdown(f"""<div class="header-gestao">
      <div>
        <p class="header-title">
          <span style="font-size:0.65rem;font-weight:500;color:rgba(255,255,255,0.45);
                       letter-spacing:3px;text-transform:uppercase;display:block;
                       margin-bottom:4px">CONNECT GROUP</span>
          Connect Valore
        </p>
        <p class="header-sub">
          TIM Corporate · {mes_str} · Bem-vindo, {name}
          <span style="display:block;margin-top:5px;font-size:0.72rem;
                       color:rgba(255,255,255,0.35);font-style:italic;letter-spacing:0.5px">
            Conexões que geram valor
          </span>
        </p>
      </div>
      <div style="display:flex;align-items:center;gap:12px">
        <img src="https://raw.githubusercontent.com/hugokhesley/dashboard-tim-connectgroup-sheets/main/logo.png"
             style="height:40px;object-fit:contain;border-radius:6px" onerror="this.style.display='none'">
        <span class="header-badge">{badge}</span>
      </div>
    </div>""", unsafe_allow_html=True)

    # ── Carrega dados ─────────────────────────────────────────────
    with st.spinner("Carregando dados..."):
        raw, bko, colab = _load_all()

    if raw.empty:
        st.warning("⚠️ Nenhum dado encontrado.")
        st.stop()

    meta_dict = dict(zip(colab["vendedor"], colab["meta"])) if not colab.empty else {}

    # ── Sidebar ───────────────────────────────────────────────────
    with st.sidebar:
        st.markdown("### 🔧 Filtros")

        meses_opcoes = gerar_meses_opcoes()
        mes_labels   = [MESES_PT.get(m[:2], m[:2]) + "/" + m[3:] for m in meses_opcoes]
        mes_idx      = st.selectbox("📅 Mês", range(len(meses_opcoes)),
                                    format_func=lambda i: mes_labels[i], index=0)
        mes_alvo = meses_opcoes[mes_idx]

        if tipo == "admin":
            parceiro_sel = st.selectbox("Parceiro / Aba", get_parceiros(raw))

            lider_opts = sorted([l for l in bko["lider"].unique() if l and l != "Sem Equipe"]) if not bko.empty else []
            lider_sel  = st.multiselect(
                "Equipe / Líder",
                options=lider_opts,
                default=lider_opts,
                placeholder="Selecione as equipes..."
            )
        else:
            parceiro_sel = "Todos"
            lider_sel    = []

        st.markdown("---")
        if st.button("🔄 Atualizar dados"):
            st.cache_data.clear()
            st.rerun()
        st.markdown("---")
        authenticator.logout("🚪 Sair", "sidebar")
        st.markdown(f"**Mês:** `{mes_alvo}`")
        st.caption("Dados via Google Sheets · cache 3 min")

    is_mes_atual = (mes_alvo == datetime.now().strftime("%m/%Y"))

    # ── Aplica filtros e merge com BKO ────────────────────────────
    df = apply_filters(raw.copy(), mes_alvo, ["NOVO", "ADITIVO"], "Todos")  # parceiro ignorado na hierarquia por lider

    # Fonte por período: mês corrente (tramitando) → só aba DadosRadar;
    # mês passado → só aba resultados. Sem isto o load_data concatena as
    # duas abas e a mesma venda soma acessos/receita em dobro.
    if "_aba" in df.columns:
        aba_norm = df["_aba"].apply(lambda x: _s(x).strip().lower())
        if is_mes_atual:
            df = df[aba_norm == "dadosradar"].copy()
        else:
            df = df[aba_norm == "resultados"].copy()

    if not bko.empty and "pedido" in df.columns:
        df["pedido"] = df["pedido"].apply(_norm_pedido)
        bk = bko.copy()
        bk["pedido"] = bk["pedido"].apply(_norm_pedido)
        df = df.merge(bk[["pedido","vendedor_real","lider"]], on="pedido", how="left")
        df["vendedor_real"] = df["vendedor_real"].apply(lambda x: _s(x) if _s(x) else "Sem Vendedor")
        df["lider"]         = df["lider"].apply(lambda x: _s(x) if _s(x) else "Sem Equipe")
    else:
        df["vendedor_real"] = "Sem Vendedor"
        df["lider"]         = "Sem Equipe"

    # Filtro por tipo de usuário
    if tipo == "lider" and lider_u:
        df = df[df["lider"].apply(lambda x: _s(x).upper()) == lider_u.upper()]
    elif tipo == "parceiro" and parceiro_u:
        if "parceiro" in df.columns:
            df = df[df["parceiro"].apply(lambda x: _s(x).upper()) == parceiro_u.upper()]
    elif tipo == "admin" and lider_sel:
        df = df[df["lider"].isin(lider_sel)]

    if not is_mes_atual:
        df = df[df["mes_ativacao"] == mes_alvo].copy()

    if df.empty:
        st.info("Nenhum dado para os filtros selecionados.")
        st.stop()

    # ── df_atv gerado DEPOIS de todos os filtros e merge ──────────
    atv = df[df["mes_ativacao"] == mes_alvo]

    # ── KPIs ──────────────────────────────────────────────────────
    ac_g   = int(atv["acessos"].sum())
    rec_g  = atv["preco_oferta"].sum()
    pip_g  = int(df[df["mes_ativacao"].isna()]["acessos"].sum())
    nv_g   = df["vendedor_real"].nunique()
    meta_g = sum(_meta_vend(v, meta_dict) for v in df["vendedor_real"].unique()) if meta_dict else nv_g * META_VENDEDOR_PAD
    pct_g  = min(int(rec_g / meta_g * 100), 999) if meta_g > 0 else 0

    k1, k2, k3, k4 = st.columns(4)
    with k1:
        st.markdown(f"""<div class="kpi-mini blue"><div class="kpi-label">🎯 Acessos Ativados</div>
          <div class="kpi-value">{ac_g:,}</div><div class="kpi-sub">no mês</div></div>""", unsafe_allow_html=True)
    with k2:
        st.markdown(f"""<div class="kpi-mini green"><div class="kpi-label">💰 Receita Ativada</div>
          <div class="kpi-value">R$ {rec_g:,.2f}</div><div class="kpi-sub">{pct_g}% da meta</div></div>""", unsafe_allow_html=True)
    with k3:
        st.markdown(f"""<div class="kpi-mini amber"><div class="kpi-label">⏳ Pipeline</div>
          <div class="kpi-value">{pip_g:,}</div><div class="kpi-sub">em tramitação</div></div>""", unsafe_allow_html=True)
    with k4:
        st.markdown(f"""<div class="kpi-mini purple"><div class="kpi-label">👥 Vendedores</div>
          <div class="kpi-value">{nv_g}</div><div class="kpi-sub">ativos</div></div>""", unsafe_allow_html=True)

    st.markdown("")

    # ── Carteira — carrega gc e dados ────────────────────────────
    gc_carteira = get_gspread_client()
    with st.spinner("Carregando carteira..."):
        df_carteira = _load_carteira(gc_carteira)
        df_carteira = _atualizar_expirados_gc(gc_carteira, df_carteira.copy())
        try:
            vends_cart, mapa_lider_cart, mapa_tbp_cart = _load_colab_carteira()
        except Exception:
            vends_cart, mapa_lider_cart, mapa_tbp_cart = [], {}, {}

    # ── Tabs ─────────────────────────────────────────────────────
    if tipo == "parceiro":
        tabs = st.tabs(["📋 Detalhado", "📊 Qualidade", "📁 Carteira", "➕ Novo Atendimento"])
        with tabs[0]:
            st.markdown('<p class="section-title">📋 Visão Detalhada por Vendedor</p>', unsafe_allow_html=True)
            render_detalhado(df, mes_alvo, meta_dict)
            st.markdown("---")
            _render_exportar(df, meta_dict, mes_alvo)
        with tabs[1]:
            _tela_qualidade(bko, info, tipo, lider_u)
        with tabs[2]:
            _tela_carteira_lista(gc_carteira, df_carteira, info, username, is_admin=(tipo=="admin"))
        with tabs[3]:
            _tela_carteira_novo(gc_carteira, df_carteira, vends_cart, mapa_lider_cart, mapa_tbp_cart, username, info)
    else:
        tab_det, tab_atr, tab_qual, tab_cart, tab_novo = st.tabs([
            "📋 Detalhado",
            "👤 Atribuição de Vendedores",
            "📊 Qualidade",
            "📁 Carteira",
            "➕ Novo Atendimento",
        ])
        with tab_det:
            st.markdown('<p class="section-title">📋 Visão Detalhada por Vendedor</p>', unsafe_allow_html=True)
            render_detalhado(df, mes_alvo, meta_dict)
            st.markdown("---")
            _render_exportar(df, meta_dict, mes_alvo)
        with tab_atr:
            st.markdown('<p class="section-title">👤 Pedidos sem Vendedor Atribuído</p>', unsafe_allow_html=True)
            with st.spinner("Carregando pendentes..."):
                df_pend, ws_bko = _carregar_pendentes_bko(
                    lider_filtro=lider_u, parceiro_filtro=parceiro_u, tipo=tipo)
                vendedores = sorted([v for v in colab["vendedor"].dropna().unique() if _s(v)]) if not colab.empty else []
            render_atribuicao(df_pend, ws_bko, vendedores)
        with tab_qual:
            _tela_qualidade(bko, info, tipo, lider_u)
        with tab_cart:
            _tela_carteira_lista(gc_carteira, df_carteira, info, username, is_admin=(tipo=="admin"))
        with tab_novo:
            _tela_carteira_novo(gc_carteira, df_carteira, vends_cart, mapa_lider_cart, mapa_tbp_cart, username, info)


# ─────────────────────────────────────────────────────────────────
#  QUALIDADE — cruzamento Safras × BKO-VENDEDOR-REAL
# ─────────────────────────────────────────────────────────────────

QUALIDADE_SPREADSHEET_ID = "1QO4EpMnVFbHMMGASzjhmemlC2avuyhZ8bHIcdSnAnlU"
ABA_QUALIDADE            = "BASE_SAFRAS_QUALIDADE"
ABA_RADAR_QUAL           = "DadosRadar"
ABA_BKO_QUAL             = "BKO-VENDEDOR-REAL"

COR_ADIM = {
    "SIM":           "#22c55e",
    "NÃO":           "#ef4444",
    "FATURA GERADA": "#f59e0b",
}



main()
